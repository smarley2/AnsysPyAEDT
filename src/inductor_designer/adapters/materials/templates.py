from __future__ import annotations

import io
from copy import copy
from dataclasses import dataclass
from importlib.resources import files

from openpyxl import load_workbook  # type: ignore[import-untyped]
from openpyxl.cell.cell import Cell  # type: ignore[import-untyped]
from openpyxl.worksheet.worksheet import Worksheet  # type: ignore[import-untyped]

from inductor_designer.domain.units import from_canonical
from inductor_designer.geometry.naming import sanitize_identifier
from inductor_designer.materials.records import MaterialRecord, PointSeries, SeriesKind


class MaterialTemplateExportError(ValueError):
    """Raised when a material record cannot be represented as an editable workbook."""


@dataclass(frozen=True, slots=True)
class MaterialTemplateDownload:
    filename: str
    content_type: str
    data: bytes


_TEMPLATES = {
    "csv": ("material-import-template.csv", "text/csv"),
    "xlsx": (
        "material-import-template.xlsx",
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    ),
}


def material_import_template(file_format: str) -> MaterialTemplateDownload:
    """Return an immutable packaged material import template."""
    try:
        filename, content_type = _TEMPLATES[file_format]
    except KeyError as error:
        raise ValueError("file_format must be 'csv' or 'xlsx'") from error
    data = files("inductor_designer.resources.material_templates").joinpath(filename).read_bytes()
    return MaterialTemplateDownload(filename, content_type, data)


def _set_cell_value(cell: Cell, value: object) -> None:
    cell.value = value
    if isinstance(value, str):
        cell.data_type = "s"


def _replace_rows(sheet: Worksheet, rows: list[tuple[object, ...]], table_name: str) -> None:
    column_count = sheet.max_column
    first_styles = tuple(
        copy(sheet.cell(row=2, column=column)._style)
        for column in range(1, column_count + 1)
    )
    middle_styles = tuple(
        copy(sheet.cell(row=3, column=column)._style)
        for column in range(1, column_count + 1)
    )
    last_styles = tuple(
        copy(sheet.cell(row=sheet.max_row, column=column)._style)
        for column in range(1, column_count + 1)
    )
    first_borders = tuple(
        copy(sheet.cell(row=2, column=column).border)
        for column in range(1, column_count + 1)
    )
    last_borders = tuple(
        copy(sheet.cell(row=sheet.max_row, column=column).border)
        for column in range(1, column_count + 1)
    )
    if sheet.max_row > 1:
        sheet.delete_rows(2, sheet.max_row - 1)
    for row_number, values in enumerate(rows, start=2):
        index = row_number - 2
        styles = (
            first_styles
            if index == 0
            else last_styles
            if index == len(rows) - 1
            else middle_styles
        )
        for column, value in enumerate(values, start=1):
            cell = sheet.cell(row=row_number, column=column)
            _set_cell_value(cell, value)
            cell._style = copy(styles[column - 1])
            if len(rows) == 1:
                border = copy(first_borders[column - 1])
                border.top = copy(last_borders[column - 1].top)
                cell.border = border
    last_row = max(1, len(rows) + 1)
    table_ref = f"A1:{sheet.cell(row=last_row, column=column_count).coordinate}"
    table = sheet.tables[table_name]
    table.ref = table_ref
    table.autoFilter.ref = table_ref
    validation_last_row = max(200, last_row)
    for validation in sheet.data_validations.dataValidation:
        first_cell = str(validation.sqref).partition(":")[0]
        validation_column = first_cell.rstrip("0123456789")
        validation.sqref = (
            f"{validation_column}2:{validation_column}{validation_last_row}"
        )


def _bh_rows(series: PointSeries) -> list[tuple[object, ...]]:
    return [
        (
            series.series_id,
            series.conditions.temperature_c,
            series.conditions.dc_bias_a_per_m,
            series.x_unit,
            series.y_unit,
            from_canonical(point.x, series.x_unit),
            from_canonical(point.y, series.y_unit),
        )
        for point in series.points
    ]


def _loss_rows(series: PointSeries) -> list[tuple[object, ...]]:
    return [
        (
            series.series_id,
            series.conditions.frequency_hz,
            series.conditions.temperature_c,
            series.conditions.dc_bias_a_per_m,
            series.x_unit,
            series.y_unit,
            from_canonical(point.x, series.x_unit),
            from_canonical(point.y, series.y_unit),
        )
        for point in series.points
    ]


def export_material_record_xlsx(
    record: MaterialRecord,
    *,
    exported_at: str | None = None,
) -> MaterialTemplateDownload:
    """Export all tabular curves in a material record as an editable workbook."""
    if not record.series:
        raise MaterialTemplateExportError(
            "Material has no curves to export; add B-H or loss curves before editing a workbook."
        )

    template = material_import_template("xlsx")
    workbook = load_workbook(io.BytesIO(template.data), data_only=False)
    metadata = {
        "manufacturer": record.ref.manufacturer,
        "material_name": record.ref.name,
        "grade": record.ref.grade,
        "source_url": "",
        "source_page": None,
        "captured_at": exported_at or record.created_at,
        "source_description": (
            f"Editable workbook derived from base material revision {record.revision_id}; "
            f"Material Studio retains {len(record.sources)} original provenance source(s)."
        ),
    }
    material = workbook["Material"]
    for row in range(2, material.max_row + 1):
        field = str(material.cell(row=row, column=1).value)
        _set_cell_value(material.cell(row=row, column=2), metadata[field])

    bh_rows = [
        row
        for series in record.series
        if series.kind is SeriesKind.BH_CURVE
        for row in _bh_rows(series)
    ]
    loss_rows = [
        row
        for series in record.series
        if series.kind is SeriesKind.LOSS_TABLE
        for row in _loss_rows(series)
    ]
    _replace_rows(workbook["B-H Curves"], bh_rows, "BHCurvesTable")
    _replace_rows(workbook["Loss Curves"], loss_rows, "LossCurvesTable")

    lineage = workbook.create_sheet("_MaterialStudio")
    lineage.sheet_state = "veryHidden"
    for row, values in enumerate(
        (
            ("field", "value"),
            ("format", "material-studio-edit"),
            ("version", 1),
            ("manufacturer", record.ref.manufacturer),
            ("material_name", record.ref.name),
            ("grade", record.ref.grade),
            ("base_revision_id", record.revision_id),
        ),
        start=1,
    ):
        for column, value in enumerate(values, start=1):
            _set_cell_value(lineage.cell(row=row, column=column), value)

    workbook.properties.creator = "PyAEDT Inductor Designer"
    workbook.properties.title = f"{record.ref.manufacturer} {record.ref.name} {record.ref.grade}"
    workbook.properties.subject = f"Editable material revision {record.revision_id}"
    workbook.properties.description = metadata["source_description"]
    stream = io.BytesIO()
    workbook.save(stream)
    ref = sanitize_identifier(
        "_".join((record.ref.manufacturer, record.ref.name, record.ref.grade))
    )
    return MaterialTemplateDownload(
        f"material-{ref}-{record.revision_id}.xlsx",
        template.content_type,
        stream.getvalue(),
    )
