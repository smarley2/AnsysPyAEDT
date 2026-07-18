from __future__ import annotations

import csv
import io
import math
from pathlib import PurePath
from zipfile import BadZipFile

from openpyxl import load_workbook  # type: ignore[import-untyped]
from openpyxl.cell.cell import Cell  # type: ignore[import-untyped]
from openpyxl.utils.exceptions import InvalidFileException  # type: ignore[import-untyped]
from openpyxl.worksheet.worksheet import Worksheet  # type: ignore[import-untyped]

from inductor_designer.application.services.material_import import MaterialImportError
from inductor_designer.application.services.material_table_import import (
    ImportedMaterialTable,
    MaterialTableMetadata,
    MaterialTableRow,
    import_material_rows,
)
from inductor_designer.materials.identity import MaterialRef
from inductor_designer.materials.records import SeriesKind, SourceKind

_CSV_COLUMNS = (
    "manufacturer",
    "material_name",
    "grade",
    "source_url",
    "source_page",
    "captured_at",
    "source_description",
    "series_id",
    "curve_kind",
    "frequency_hz",
    "temperature_c",
    "dc_bias_a_per_m",
    "x_unit",
    "y_unit",
    "x",
    "y",
)
_METADATA_COLUMNS = _CSV_COLUMNS[:7]
_SERIES_METADATA_COLUMNS = _CSV_COLUMNS[8:14]
_BH_UNITS = (frozenset({"A/m", "kA/m", "Oe"}), frozenset({"T", "mT", "G", "kG"}))
_LOSS_UNITS = (
    frozenset({"T", "mT", "G", "kG"}),
    frozenset({"W/m3", "kW/m3", "mW/cm3"}),
)
_VISIBLE_SHEETS = ("Instructions", "Material", "B-H Curves", "Loss Curves")
_MATERIAL_KEYS = _METADATA_COLUMNS
_BH_COLUMNS = (
    "series_id",
    "temperature_c",
    "dc_bias_a_per_m",
    "h_unit",
    "b_unit",
    "h",
    "b",
)
_LOSS_COLUMNS = (
    "series_id",
    "frequency_hz",
    "temperature_c",
    "dc_bias_a_per_m",
    "b_unit",
    "loss_unit",
    "b",
    "loss",
)


def _fail(filename: str, location: str, message: str) -> MaterialImportError:
    return MaterialImportError((f"{filename}: {location}: {message}",))


def _required_text(value: object, filename: str, location: str, field: str) -> str:
    if not isinstance(value, str) or not value.strip():
        raise _fail(filename, location, f"{field} must not be blank")
    return value


def _number(
    value: object,
    filename: str,
    location: str,
    field: str,
    *,
    optional: bool = False,
) -> float | None:
    if optional and (value is None or value == ""):
        return None
    if isinstance(value, bool):
        raise _fail(filename, location, f"{field} must be a finite number")
    try:
        parsed = float(value)  # type: ignore[arg-type]
    except (TypeError, ValueError) as error:
        raise _fail(filename, location, f"{field} must be a finite number") from error
    if not math.isfinite(parsed):
        raise _fail(filename, location, f"{field} must be a finite number")
    return parsed


def _optional_page(value: object, filename: str, location: str) -> int | None:
    if value is None or value == "":
        return None
    number = _number(value, filename, location, "source_page")
    if number is None or not number.is_integer():
        raise _fail(filename, location, "source_page must be an integer")
    return int(number)


def _kind(value: object, filename: str, location: str) -> SeriesKind:
    try:
        return SeriesKind(value)
    except (TypeError, ValueError) as error:
        allowed = ", ".join(kind.value for kind in SeriesKind)
        raise _fail(filename, location, f"curve_kind must be one of: {allowed}") from error


def _validate_units(
    kind: SeriesKind,
    x_unit: str,
    y_unit: str,
    filename: str,
    location: str,
) -> None:
    x_units, y_units = _BH_UNITS if kind is SeriesKind.BH_CURVE else _LOSS_UNITS
    if x_unit not in x_units:
        raise _fail(filename, location, f"x_unit '{x_unit}' is not valid for {kind.value}")
    if y_unit not in y_units:
        raise _fail(filename, location, f"y_unit '{y_unit}' is not valid for {kind.value}")


def _metadata(values: dict[str, object], filename: str, location: str) -> MaterialTableMetadata:
    manufacturer = _required_text(values["manufacturer"], filename, location, "manufacturer")
    material_name = _required_text(values["material_name"], filename, location, "material_name")
    grade = _required_text(values["grade"], filename, location, "grade")
    captured_at = _required_text(values["captured_at"], filename, location, "captured_at")
    description = _required_text(
        values["source_description"], filename, location, "source_description"
    )
    source_url = values["source_url"]
    if not isinstance(source_url, str):
        raise _fail(filename, location, "source_url must be text")
    return MaterialTableMetadata(
        ref=MaterialRef(manufacturer, material_name, grade),
        source_url=source_url,
        source_page=_optional_page(values["source_page"], filename, location),
        captured_at=captured_at,
        source_description=description,
    )


def _row(values: dict[str, object], filename: str, location: str) -> MaterialTableRow:
    kind = _kind(values["curve_kind"], filename, location)
    series_id = _required_text(values["series_id"], filename, location, "series_id")
    x_unit = _required_text(values["x_unit"], filename, location, "x_unit")
    y_unit = _required_text(values["y_unit"], filename, location, "y_unit")
    _validate_units(kind, x_unit, y_unit, filename, location)
    frequency = _number(values["frequency_hz"], filename, location, "frequency_hz", optional=True)
    if kind is SeriesKind.LOSS_TABLE and (frequency is None or frequency <= 0):
        raise _fail(filename, location, "frequency_hz must be positive for loss-table")
    return MaterialTableRow(
        series_id=series_id,
        kind=kind,
        frequency_hz=frequency,
        temperature_c=_number(
            values["temperature_c"], filename, location, "temperature_c", optional=True
        ),
        dc_bias_a_per_m=_number(
            values["dc_bias_a_per_m"], filename, location, "dc_bias_a_per_m", optional=True
        ),
        x_unit=x_unit,
        y_unit=y_unit,
        x=_number(values["x"], filename, location, "x") or 0.0,
        y=_number(values["y"], filename, location, "y") or 0.0,
    )


def _import_csv(filename: str, data: bytes) -> ImportedMaterialTable:
    try:
        text = data.decode("utf-8-sig")
    except UnicodeDecodeError as error:
        raise _fail(filename, "CSV", "file must be valid UTF-8") from error
    try:
        reader = csv.DictReader(io.StringIO(text), strict=True)
        header = tuple(reader.fieldnames or ())
        if header != _CSV_COLUMNS:
            raise _fail(
                filename,
                "CSV header",
                f"expected exactly these columns: {', '.join(_CSV_COLUMNS)}",
            )
        values = list(reader)
    except csv.Error as error:
        raise _fail(filename, f"CSV row {getattr(reader, 'line_num', 1)}", str(error)) from error

    if not values:
        raise _fail(filename, "CSV row 2", "material table requires at least one data row")

    first_metadata = tuple(values[0][field] for field in _METADATA_COLUMNS)
    series_metadata: dict[str, tuple[object, ...]] = {}
    rows: list[MaterialTableRow] = []
    for row_number, values_by_name in enumerate(values, start=2):
        location = f"CSV row {row_number}"
        if None in values_by_name:
            raise _fail(filename, location, f"expected exactly {len(_CSV_COLUMNS)} values")
        if tuple(values_by_name[field] for field in _METADATA_COLUMNS) != first_metadata:
            raise _fail(filename, location, "material and source metadata must match the first row")
        parsed = _row(values_by_name, filename, location)
        repeated = tuple(values_by_name[field] for field in _SERIES_METADATA_COLUMNS)
        if previous := series_metadata.get(parsed.series_id):
            if repeated != previous:
                raise _fail(
                    filename,
                    location,
                    f"series '{parsed.series_id}' metadata is inconsistent",
                )
        else:
            series_metadata[parsed.series_id] = repeated
        rows.append(parsed)

    metadata = _metadata(values[0], filename, "CSV row 2")
    try:
        return import_material_rows(
            metadata,
            tuple(rows),
            upload_filename=filename,
            upload_kind=SourceKind.CSV,
            upload_bytes=data,
        )
    except MaterialImportError as error:
        issues = tuple(f"{filename}: {issue}" for issue in error.issues)
        raise MaterialImportError(issues) from error


def _reject_formulas(filename: str, sheet: Worksheet, max_column: int) -> None:
    for row in sheet.iter_rows(min_row=1, max_row=sheet.max_row, max_col=max_column):
        for cell in row:
            if cell.data_type == "f":
                raise _fail(
                    filename,
                    f"{sheet.title}!{cell.coordinate}",
                    "formulas are not allowed",
                )


def _check_header(filename: str, sheet: Worksheet, expected: tuple[str, ...]) -> tuple[Cell, ...]:
    cells = tuple(sheet[1][: len(expected)])
    if tuple(cell.value for cell in cells) != expected or sheet.max_column != len(expected):
        raise _fail(
            filename,
            f"{sheet.title}!A1",
            f"expected exactly these columns: {', '.join(expected)}",
        )
    return cells


def _material_metadata(filename: str, sheet: Worksheet) -> MaterialTableMetadata:
    _reject_formulas(filename, sheet, sheet.max_column)
    if sheet.max_column != 2:
        location = "Material!A1"
        if sheet.max_column > 2:
            extra_cell = next(
                (
                    cell
                    for row in sheet.iter_rows(min_col=3, max_col=sheet.max_column)
                    for cell in row
                    if cell.value is not None or cell.has_style
                ),
                sheet.cell(row=1, column=3),
            )
            location = f"Material!{extra_cell.coordinate}"
        raise _fail(filename, location, "expected exactly two columns: field, value")
    if (sheet["A1"].value, sheet["B1"].value) != ("field", "value"):
        raise _fail(filename, "Material!A1", "expected columns: field, value")

    values: dict[str, object] = {}
    locations: dict[str, str] = {}
    for row_number in range(2, sheet.max_row + 1):
        key = sheet.cell(row=row_number, column=1).value
        value = sheet.cell(row=row_number, column=2).value
        if key is None and value is None:
            continue
        if not isinstance(key, str) or key not in _MATERIAL_KEYS:
            raise _fail(filename, f"Material!A{row_number}", f"unknown material key: {key!r}")
        if key in values:
            raise _fail(filename, f"Material!A{row_number}", f"duplicate material key: {key}")
        values[key] = value
        locations[key] = f"Material!B{row_number}"
    if set(values) != set(_MATERIAL_KEYS):
        missing = ", ".join(key for key in _MATERIAL_KEYS if key not in values)
        raise _fail(filename, "Material!A2", f"missing material keys: {missing}")

    manufacturer = _required_text(
        values["manufacturer"], filename, locations["manufacturer"], "manufacturer"
    )
    material_name = _required_text(
        values["material_name"], filename, locations["material_name"], "material_name"
    )
    grade = _required_text(values["grade"], filename, locations["grade"], "grade")
    source_url = values["source_url"]
    if source_url is None:
        source_url = ""
    if not isinstance(source_url, str):
        raise _fail(filename, locations["source_url"], "source_url must be text")
    return MaterialTableMetadata(
        ref=MaterialRef(manufacturer, material_name, grade),
        source_url=source_url,
        source_page=_optional_page(values["source_page"], filename, locations["source_page"]),
        captured_at=_required_text(
            values["captured_at"], filename, locations["captured_at"], "captured_at"
        ),
        source_description=_required_text(
            values["source_description"],
            filename,
            locations["source_description"],
            "source_description",
        ),
    )


def _table_rows(
    filename: str,
    sheet: Worksheet,
    columns: tuple[str, ...],
    kind: SeriesKind,
) -> list[MaterialTableRow]:
    _reject_formulas(filename, sheet, len(columns))
    _check_header(filename, sheet, columns)
    rows: list[MaterialTableRow] = []
    series_metadata: dict[str, tuple[object, ...]] = {}
    for cells in sheet.iter_rows(min_row=2, max_col=len(columns)):
        if all(cell.value is None for cell in cells):
            continue
        row_number = cells[0].row
        location = f"{sheet.title}!A{row_number}:{cells[-1].coordinate}"
        raw = dict(zip(columns, (cell.value for cell in cells), strict=True))
        if kind is SeriesKind.BH_CURVE:
            values: dict[str, object] = {
                "series_id": raw["series_id"],
                "curve_kind": kind.value,
                "frequency_hz": None,
                "temperature_c": raw["temperature_c"],
                "dc_bias_a_per_m": raw["dc_bias_a_per_m"],
                "x_unit": raw["h_unit"],
                "y_unit": raw["b_unit"],
                "x": raw["h"],
                "y": raw["b"],
            }
        else:
            values = {
                "series_id": raw["series_id"],
                "curve_kind": kind.value,
                "frequency_hz": raw["frequency_hz"],
                "temperature_c": raw["temperature_c"],
                "dc_bias_a_per_m": raw["dc_bias_a_per_m"],
                "x_unit": raw["b_unit"],
                "y_unit": raw["loss_unit"],
                "x": raw["b"],
                "y": raw["loss"],
            }
        parsed = _row(values, filename, location)
        repeated = (
            parsed.kind,
            parsed.frequency_hz,
            parsed.temperature_c,
            parsed.dc_bias_a_per_m,
            parsed.x_unit,
            parsed.y_unit,
        )
        if previous := series_metadata.get(parsed.series_id):
            if repeated != previous:
                raise _fail(
                    filename,
                    location,
                    f"series '{parsed.series_id}' metadata is inconsistent",
                )
        else:
            series_metadata[parsed.series_id] = repeated
        rows.append(parsed)
    return rows


def _import_xlsx(filename: str, data: bytes) -> ImportedMaterialTable:
    try:
        workbook = load_workbook(io.BytesIO(data), read_only=False, data_only=False)
    except (BadZipFile, InvalidFileException, OSError, ValueError) as error:
        raise _fail(filename, "workbook", "file is not a valid .xlsx workbook") from error

    visible = frozenset(
        sheet.title for sheet in workbook.worksheets if sheet.sheet_state == "visible"
    )
    extras = tuple(sheet for sheet in workbook.worksheets if sheet.title not in _VISIBLE_SHEETS)
    valid_lists = not extras or (
        len(extras) == 1 and extras[0].title == "_Lists" and extras[0].sheet_state != "visible"
    )
    if visible != frozenset(_VISIBLE_SHEETS) or not valid_lists:
        raise _fail(
            filename,
            "workbook sheets",
            "expected exactly four visible sheets: " + ", ".join(_VISIBLE_SHEETS),
        )

    metadata = _material_metadata(filename, workbook["Material"])
    rows = _table_rows(filename, workbook["B-H Curves"], _BH_COLUMNS, SeriesKind.BH_CURVE)
    rows.extend(
        _table_rows(filename, workbook["Loss Curves"], _LOSS_COLUMNS, SeriesKind.LOSS_TABLE)
    )
    if not rows:
        raise _fail(filename, "B-H Curves!A2 / Loss Curves!A2", "requires at least one data row")
    try:
        return import_material_rows(
            metadata,
            tuple(rows),
            upload_filename=filename,
            upload_kind=SourceKind.SPREADSHEET,
            upload_bytes=data,
        )
    except MaterialImportError as error:
        issues = tuple(f"{filename}: {issue}" for issue in error.issues)
        raise MaterialImportError(issues) from error


def import_material_file(filename: str, data: bytes) -> ImportedMaterialTable:
    """Decode a CSV or Excel upload and import all contained series."""
    extension = PurePath(filename).suffix
    if extension == ".csv":
        return _import_csv(filename, data)
    if extension == ".xlsx":
        return _import_xlsx(filename, data)
    raise _fail(filename, "extension", "expected .csv or .xlsx")
