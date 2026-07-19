from __future__ import annotations

import io
from dataclasses import replace

import pytest
from openpyxl import load_workbook

from inductor_designer.adapters.materials.table_file import (
    import_material_file,
    import_material_file_as_draft,
)
from inductor_designer.adapters.materials.templates import (
    MaterialTemplateExportError,
    export_material_record_xlsx,
    material_import_template,
)
from inductor_designer.materials.identity import MaterialRef
from inductor_designer.materials.records import (
    CurveConditions,
    CurvePoint,
    MaterialRecord,
    MaterialStatus,
    PointSeries,
    SeriesKind,
    SourceKind,
    SourceProvenance,
)


def _approved_record() -> MaterialRecord:
    source = SourceProvenance(
        kind=SourceKind.CSV,
        filename="original.csv",
        sha256="a" * 64,
        url="https://example.com/material",
        page=7,
        captured_at="2026-07-18T12:00:00+00:00",
        description="Original source",
    )
    series = (
        PointSeries(
            "bh-25c",
            SeriesKind.BH_CURVE,
            "Oe",
            "kG",
            CurveConditions(None, 25.0, None),
            (CurvePoint(0.0, 0.0), CurvePoint(79.577471546, 0.1)),
            source.filename,
            None,
        ),
        PointSeries(
            "bh-100c",
            SeriesKind.BH_CURVE,
            "kA/m",
            "mT",
            CurveConditions(None, 100.0, 50.0),
            (CurvePoint(1000.0, 100.0e-3), CurvePoint(2000.0, 200.0e-3)),
            source.filename,
            None,
        ),
        PointSeries(
            "loss-100khz",
            SeriesKind.LOSS_TABLE,
            "kG",
            "mW/cm3",
            CurveConditions(100_000.0, 25.0, 0.0),
            (CurvePoint(0.05, 1000.0), CurvePoint(0.1, 2000.0)),
            source.filename,
            None,
        ),
        PointSeries(
            "loss-200khz",
            SeriesKind.LOSS_TABLE,
            "mT",
            "kW/m3",
            CurveConditions(200_000.0, 80.0, 10.0),
            (CurvePoint(0.05, 3000.0), CurvePoint(0.1, 6000.0)),
            source.filename,
            None,
        ),
    )
    return MaterialRecord(
        ref=MaterialRef("Example Magnetics", "Ferrite", "F1"),
        revision_id="abcde1234567",
        status=MaterialStatus.APPROVED,
        created_at="2026-07-18T12:00:00+00:00",
        reviewed_by="reviewer@example.com",
        approved_by="approver@example.com",
        sources=(source,),
        series=series,
        relative_permeability=2000.0,
        steinmetz=None,
        notes="Approved base",
    )


def test_packaged_templates_have_download_metadata_and_equivalent_examples() -> None:
    csv_download = material_import_template("csv")
    xlsx_download = material_import_template("xlsx")

    assert csv_download.filename == "material-import-template.csv"
    assert csv_download.content_type == "text/csv"
    assert isinstance(csv_download.data, bytes)
    assert csv_download.data
    assert xlsx_download.filename == "material-import-template.xlsx"
    assert (
        xlsx_download.content_type
        == "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    assert isinstance(xlsx_download.data, bytes)
    assert xlsx_download.data

    csv_result = import_material_file(csv_download.filename, csv_download.data)
    xlsx_result = import_material_file(xlsx_download.filename, xlsx_download.data)
    assert xlsx_result.ref == csv_result.ref
    assert xlsx_result.series == csv_result.series


def test_exported_workbook_carries_hidden_base_identity_but_template_does_not() -> None:
    record = _approved_record()

    exported = export_material_record_xlsx(record)
    workbook = load_workbook(io.BytesIO(exported.data), data_only=False)
    lineage = workbook["_MaterialStudio"]

    assert lineage.sheet_state == "veryHidden"
    assert tuple(lineage.values) == (
        ("field", "value"),
        ("format", "material-studio-edit"),
        ("version", 1),
        ("manufacturer", record.ref.manufacturer),
        ("material_name", record.ref.name),
        ("grade", record.ref.grade),
        ("base_revision_id", record.revision_id),
    )
    imported = import_material_file_as_draft(
        exported.filename,
        exported.data,
        created_at="2026-07-19T10:00:00+00:00",
    )
    assert imported.base_ref == record.ref
    assert imported.base_revision_id == record.revision_id

    template = material_import_template("xlsx")
    generic = import_material_file_as_draft(
        template.filename,
        template.data,
        created_at="2026-07-19T10:00:00+00:00",
    )
    assert generic.base_ref is None
    assert generic.base_revision_id is None


def test_packaged_xlsx_has_documented_structure_and_validations() -> None:
    download = material_import_template("xlsx")
    workbook = load_workbook(io.BytesIO(download.data), data_only=False)

    assert tuple(sheet.title for sheet in workbook.worksheets) == (
        "Instructions",
        "Material",
        "B-H Curves",
        "Loss Curves",
    )
    assert all(sheet.sheet_state == "visible" for sheet in workbook.worksheets)

    material = workbook["Material"]
    bh = workbook["B-H Curves"]
    loss = workbook["Loss Curves"]
    assert tuple(cell.value for cell in material[1]) == ("field", "value")
    assert tuple(cell.value for cell in bh[1]) == (
        "series_id",
        "temperature_c",
        "dc_bias_a_per_m",
        "h_unit",
        "b_unit",
        "h",
        "b",
    )
    assert tuple(cell.value for cell in loss[1]) == (
        "series_id",
        "frequency_hz",
        "temperature_c",
        "dc_bias_a_per_m",
        "b_unit",
        "loss_unit",
        "b",
        "loss",
    )
    for sheet in (material, bh, loss):
        assert tuple(table.ref for table in sheet.tables.values()) == (sheet.dimensions,)

    validations = {
        (sheet.title, validation.sqref): validation.formula1
        for sheet in (bh, loss)
        for validation in sheet.data_validations.dataValidation
    }
    assert validations == {
        ("B-H Curves", "D2:D200"): '\"A/m,kA/m,Oe\"',
        ("B-H Curves", "E2:E200"): '\"T,mT,G,kG\"',
        ("Loss Curves", "E2:E200"): '\"T,mT,G,kG\"',
        ("Loss Curves", "F2:F200"): '\"W/m3,kW/m3,mW/cm3\"',
    }


def test_material_import_template_rejects_unknown_format() -> None:
    with pytest.raises(ValueError, match="csv.*xlsx"):
        material_import_template("xls")


def test_export_material_record_xlsx_preserves_template_and_retained_units() -> None:
    packaged = load_workbook(io.BytesIO(material_import_template("xlsx").data))
    download = export_material_record_xlsx(_approved_record())

    assert download.filename == "material-Example_Magnetics_Ferrite_F1-abcde1234567.xlsx"
    assert (
        download.content_type
        == "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    workbook = load_workbook(io.BytesIO(download.data), data_only=False)
    assert tuple(sheet.title for sheet in workbook.worksheets[:-1]) == tuple(
        sheet.title for sheet in packaged.worksheets
    )
    assert workbook.worksheets[-1].title == "_MaterialStudio"
    assert workbook.worksheets[-1].sheet_state == "veryHidden"
    assert workbook["B-H Curves"][1][0].style_id == packaged["B-H Curves"][1][0].style_id
    assert workbook["B-H Curves"][2][0].style_id == packaged["B-H Curves"][2][0].style_id
    assert tuple(workbook["B-H Curves"].tables) == ("BHCurvesTable",)
    assert tuple(workbook["Loss Curves"].tables) == ("LossCurvesTable",)
    assert tuple(workbook["B-H Curves"].data_validations.dataValidation) != ()
    assert tuple(workbook["Loss Curves"].data_validations.dataValidation) != ()

    material = dict(workbook["Material"].iter_rows(min_row=2, values_only=True))
    assert material == {
        "manufacturer": "Example Magnetics",
        "material_name": "Ferrite",
        "grade": "F1",
        "source_url": None,
        "source_page": None,
        "captured_at": "2026-07-18T12:00:00+00:00",
        "source_description": (
            "Editable workbook derived from base material revision abcde1234567; "
            "Material Studio retains 1 original provenance source(s)."
        ),
    }
    assert tuple(workbook["B-H Curves"].iter_rows(min_row=2, values_only=True)) == (
        ("bh-25c", 25.0, None, "Oe", "kG", 0.0, 0.0),
        ("bh-25c", 25.0, None, "Oe", "kG", pytest.approx(1.0), 1.0),
        ("bh-100c", 100.0, 50.0, "kA/m", "mT", 1.0, 100.0),
        ("bh-100c", 100.0, 50.0, "kA/m", "mT", 2.0, 200.0),
    )
    assert tuple(workbook["Loss Curves"].iter_rows(min_row=2, values_only=True)) == (
        ("loss-100khz", 100_000.0, 25.0, 0.0, "kG", "mW/cm3", 0.5, 1.0),
        ("loss-100khz", 100_000.0, 25.0, 0.0, "kG", "mW/cm3", 1.0, 2.0),
        ("loss-200khz", 200_000.0, 80.0, 10.0, "mT", "kW/m3", 50.0, 3.0),
        ("loss-200khz", 200_000.0, 80.0, 10.0, "mT", "kW/m3", 100.0, 6.0),
    )


def test_export_material_record_xlsx_rejects_scalar_only_record() -> None:
    record = replace(_approved_record(), series=())

    with pytest.raises(MaterialTemplateExportError, match="curve.*B-H.*loss"):
        export_material_record_xlsx(record)


def test_export_extends_ranges_and_preserves_first_middle_last_style_roles() -> None:
    record = _approved_record()
    many_points = tuple(
        CurvePoint(index * 79.57747154594767, index * 0.001) for index in range(205)
    )
    record = replace(record, series=(replace(record.series[0], points=many_points),))
    packaged = load_workbook(io.BytesIO(material_import_template("xlsx").data))

    workbook = load_workbook(io.BytesIO(export_material_record_xlsx(record).data))

    bh = workbook["B-H Curves"]
    loss = workbook["Loss Curves"]
    assert bh.tables["BHCurvesTable"].ref == "A1:G206"
    assert bh.tables["BHCurvesTable"].autoFilter.ref == "A1:G206"
    assert loss.tables["LossCurvesTable"].ref == "A1:H1"
    assert loss.tables["LossCurvesTable"].autoFilter.ref == "A1:H1"
    assert {
        str(validation.sqref) for validation in bh.data_validations.dataValidation
    } == {"D2:D206", "E2:E206"}
    assert {
        str(validation.sqref) for validation in loss.data_validations.dataValidation
    } == {"E2:E200", "F2:F200"}
    assert [cell.style_id for cell in bh[2]] == [
        cell.style_id for cell in packaged["B-H Curves"][2]
    ]
    assert [cell.style_id for cell in bh[3]] == [
        cell.style_id for cell in packaged["B-H Curves"][3]
    ]
    assert [cell.style_id for cell in bh[206]] == [
        cell.style_id for cell in packaged["B-H Curves"][4]
    ]


def test_export_single_row_combines_first_and_last_boundaries() -> None:
    record = _approved_record()
    record = replace(
        record,
        series=(replace(record.series[0], points=(record.series[0].points[0],)),),
    )
    packaged = load_workbook(io.BytesIO(material_import_template("xlsx").data))

    workbook = load_workbook(io.BytesIO(export_material_record_xlsx(record).data))

    exported = workbook["B-H Curves"]
    template = packaged["B-H Curves"]
    for column in range(1, exported.max_column + 1):
        actual = exported.cell(2, column)
        first = template.cell(2, column)
        last = template.cell(template.max_row, column)
        assert actual._style.fontId == first._style.fontId
        assert actual._style.fillId == first._style.fillId
        assert actual._style.alignmentId == first._style.alignmentId
        assert actual._style.numFmtId == first._style.numFmtId
        assert actual.border.top.style == last.border.top.style
        assert actual.border.top.color.rgb == last.border.top.color.rgb
        assert actual.border.bottom.style == first.border.bottom.style
        assert actual.border.bottom.color.rgb == first.border.bottom.color.rgb


def test_export_preserves_formula_like_text_as_literal_strings_on_round_trip() -> None:
    base = _approved_record()
    source = replace(
        base.sources[0],
        url="@https://example.com/material",
        captured_at="-2026-07-18T12:00:00+00:00",
    )
    record = replace(
        base,
        ref=MaterialRef("=2+2", "+Ferrite", "-F1"),
        sources=(source,),
        series=(replace(base.series[0], series_id="=bh"),),
    )

    download = export_material_record_xlsx(
        record,
        exported_at="-2026-07-19T10:00:00+00:00",
    )
    workbook = load_workbook(io.BytesIO(download.data), data_only=False)

    for coordinate in ("B2", "B3", "B4", "B7", "B8"):
        assert workbook["Material"][coordinate].data_type == "s"
    for coordinate in ("A2", "D2", "E2"):
        assert workbook["B-H Curves"][coordinate].data_type == "s"

    imported = import_material_file(download.filename, download.data)
    assert imported.ref == record.ref
    assert imported.sources[0].url == ""
    assert imported.sources[0].page is None
    assert imported.sources[0].captured_at == "-2026-07-19T10:00:00+00:00"
    assert "base material revision" in imported.sources[0].description
    assert imported.series[0].series_id == "=bh"
