from __future__ import annotations

import csv
import io
from copy import deepcopy
from datetime import datetime

import pytest
from openpyxl import Workbook, load_workbook

from inductor_designer.adapters.materials.table_file import (
    import_material_file,
    import_material_file_as_draft,
    import_material_file_as_imported,
)
from inductor_designer.adapters.materials.templates import (
    export_material_record_xlsx,
    material_import_template,
)
from inductor_designer.application.services.material_import import MaterialImportError
from inductor_designer.materials.identity import MaterialRef
from inductor_designer.materials.records import (
    CurveConditions,
    CurvePoint,
    MaterialRecord,
    MaterialStatus,
    PointSeries,
    SeriesKind,
    SourceKind,
    SourceProvenance,
)

CSV_COLUMNS = (
    "manufacturer",
    "material_name",
    "grade",
    "source_url",
    "source_page",
    "captured_at",
    "source_description",
    "series_id",
    "curve_kind",
    "frequency_hz",
    "temperature_c",
    "dc_bias_a_per_m",
    "x_unit",
    "y_unit",
    "x",
    "y",
)


def _csv_bytes(rows: list[dict[str, object]]) -> bytes:
    stream = io.StringIO(newline="")
    writer = csv.DictWriter(stream, fieldnames=CSV_COLUMNS, lineterminator="\n")
    writer.writeheader()
    writer.writerows(rows)
    return stream.getvalue().encode()


def _row(**changes: object) -> dict[str, object]:
    row: dict[str, object] = {
        "manufacturer": "Example",
        "material_name": "Ferrite",
        "grade": "F1",
        "source_url": "https://example.com/f1",
        "source_page": 7,
        "captured_at": "2026-07-18T12:00:00+00:00",
        "source_description": "Example material table",
        "series_id": "bh-25c",
        "curve_kind": "bh-curve",
        "frequency_hz": "",
        "temperature_c": 25,
        "dc_bias_a_per_m": "",
        "x_unit": "Oe",
        "y_unit": "kG",
        "x": 0,
        "y": 0,
    }
    row.update(changes)
    return row


def _workbook_bytes(
    *,
    formula_cell: tuple[str, str] | None = None,
    cell_value: tuple[str, str, object] | None = None,
    remove_sheet: str | None = None,
    empty_tables: bool = False,
    remove_material_value_column: bool = False,
    reorder_sheets: bool = False,
) -> bytes:
    workbook = Workbook()
    instructions = workbook.active
    instructions.title = "Instructions"
    instructions["A1"] = "Replace the synthetic example rows before review."

    material = workbook.create_sheet("Material")
    material.append(("field", "value"))
    for item in (
        ("manufacturer", "Example"),
        ("material_name", "Ferrite"),
        ("grade", "F1"),
        ("source_url", "https://example.com/f1"),
        ("source_page", 7),
        ("captured_at", "2026-07-18T12:00:00+00:00"),
        ("source_description", "Example material table"),
    ):
        material.append(item)

    bh = workbook.create_sheet("B-H Curves")
    bh.append(
        (
            "series_id",
            "temperature_c",
            "dc_bias_a_per_m",
            "h",
            "h_unit",
            "b",
            "b_unit",
        )
    )
    bh.append(("bh-25c", 25, None, 0, "Oe", 0, "kG"))
    bh.append(("bh-25c", 25, None, 1, "Oe", 1, "kG"))

    loss = workbook.create_sheet("Loss Curves")
    loss.append(
        (
            "series_id",
            "frequency_hz",
            "temperature_c",
            "dc_bias_a_per_m",
            "b",
            "b_unit",
            "loss",
            "loss_unit",
        )
    )
    loss.append(("loss-100khz", 100_000, None, 0, 1, "kG", 2, "mW/cm3"))
    loss.append(("loss-100khz", 100_000, None, 0, 2, "kG", 4, "mW/cm3"))

    if empty_tables:
        bh.delete_rows(2, bh.max_row)
        loss.delete_rows(2, loss.max_row)
    if remove_material_value_column:
        material.delete_cols(2)

    if formula_cell is not None:
        sheet, coordinate = formula_cell
        workbook[sheet][coordinate] = "=1+1"
    if cell_value is not None:
        sheet, coordinate, value = cell_value
        workbook[sheet][coordinate] = value
    if remove_sheet is not None:
        del workbook[remove_sheet]
    if reorder_sheets:
        workbook.move_sheet(loss, offset=-3)

    stream = io.BytesIO()
    workbook.save(stream)
    return stream.getvalue()


def test_import_material_csv_groups_series_and_preserves_metadata() -> None:
    data = _csv_bytes(
        [
            _row(),
            _row(x=1, y=1),
            _row(
                series_id="loss-100khz",
                curve_kind="loss-table",
                frequency_hz=100_000,
                temperature_c="",
                dc_bias_a_per_m=0,
                x_unit="kG",
                y_unit="mW/cm3",
                x=1,
                y=2,
            ),
            _row(
                series_id="loss-100khz",
                curve_kind="loss-table",
                frequency_hz=100_000,
                temperature_c="",
                dc_bias_a_per_m=0,
                x_unit="kG",
                y_unit="mW/cm3",
                x=2,
                y=4,
            ),
        ]
    )

    result = import_material_file("example.csv", data)

    assert result.ref == MaterialRef("Example", "Ferrite", "F1")
    assert tuple(item.kind for item in result.series) == (
        SeriesKind.BH_CURVE,
        SeriesKind.LOSS_TABLE,
    )
    assert result.series[0].conditions == CurveConditions(None, 25.0, None)
    assert result.series[0].points == (
        CurvePoint(0.0, 0.0),
        CurvePoint(79.577471546, 0.1),
    )
    assert result.series[1].conditions == CurveConditions(100_000.0, None, 0.0)
    assert result.series[1].points == (
        CurvePoint(0.0, 0.0),
        CurvePoint(0.1, 2000.0),
        CurvePoint(0.2, 4000.0),
    )
    assert result.sources[0].kind is SourceKind.CSV
    assert result.sources[0].filename == "example.csv"
    assert result.sources[0].url == "https://example.com/f1"
    assert result.sources[0].page == 7
    assert result.sources[0].captured_at == "2026-07-18T12:00:00+00:00"
    assert result.sources[0].description == "Example material table"


def test_equivalent_csv_and_xlsx_produce_equal_ref_and_series() -> None:
    csv_data = _csv_bytes(
        [
            _row(),
            _row(x=1, y=1),
            _row(
                series_id="loss-100khz",
                curve_kind="loss-table",
                frequency_hz=100_000,
                temperature_c="",
                dc_bias_a_per_m=0,
                x_unit="kG",
                y_unit="mW/cm3",
                x=1,
                y=2,
            ),
            _row(
                series_id="loss-100khz",
                curve_kind="loss-table",
                frequency_hz=100_000,
                temperature_c="",
                dc_bias_a_per_m=0,
                x_unit="kG",
                y_unit="mW/cm3",
                x=2,
                y=4,
            ),
        ]
    )

    csv_result = import_material_file("example.csv", csv_data)
    xlsx_result = import_material_file("example.xlsx", _workbook_bytes())

    assert xlsx_result.ref == csv_result.ref
    assert xlsx_result.series == csv_result.series
    assert xlsx_result.sources[0].kind is SourceKind.SPREADSHEET


def test_xlsx_accepts_documented_sheets_in_any_tab_order() -> None:
    result = import_material_file("reordered.xlsx", _workbook_bytes(reorder_sheets=True))

    assert result.ref == MaterialRef("Example", "Ferrite", "F1")


def test_xlsx_normalizes_excel_date_for_captured_at() -> None:
    result = import_material_file(
        "excel-date.xlsx",
        _workbook_bytes(cell_value=("Material", "B7", datetime(2026, 7, 23))),
    )

    assert result.sources[0].captured_at == "2026-07-23T00:00:00"


def test_xlsx_rejects_missing_material_value_column() -> None:
    _assert_import_error(
        "missing-material-column.xlsx",
        _workbook_bytes(remove_material_value_column=True),
        "Material!A1",
        "field, value",
    )


@pytest.mark.parametrize(
    ("value", "expected"),
    [("unexpected", "exactly two columns"), ("=1+1", "formulas are not allowed")],
)
def test_xlsx_rejects_extra_material_column(value: str, expected: str) -> None:
    _assert_import_error(
        "extra-material-column.xlsx",
        _workbook_bytes(cell_value=("Material", "C2", value)),
        "Material!C2",
        expected,
    )


@pytest.mark.parametrize(
    ("sheet", "coordinate"),
    [("Material", "B2"), ("B-H Curves", "D2"), ("Loss Curves", "G2")],
)
def test_xlsx_rejects_formulas_with_sheet_and_cell(sheet: str, coordinate: str) -> None:
    with pytest.raises(MaterialImportError) as caught:
        import_material_file("formula.xlsx", _workbook_bytes(formula_cell=(sheet, coordinate)))

    assert "formula.xlsx" in str(caught.value)
    assert sheet in str(caught.value)
    assert coordinate in str(caught.value)
    assert "formula" in str(caught.value).lower()


def _assert_import_error(filename: str, data: bytes, *messages: str) -> None:
    with pytest.raises(MaterialImportError) as caught:
        import_material_file(filename, data)

    text = str(caught.value)
    assert filename in text
    assert all(message in text for message in messages)


def test_rejects_unsupported_or_uppercase_extension() -> None:
    _assert_import_error("material.xls", b"", ".csv", ".xlsx")
    _assert_import_error("material.CSV", b"", ".csv", ".xlsx")


def test_csv_rejects_invalid_utf8() -> None:
    _assert_import_error("invalid.csv", b"\xff", "UTF-8")


def test_csv_rejects_malformed_quoting() -> None:
    data = (",".join(CSV_COLUMNS) + '\n"unterminated').encode()

    _assert_import_error("malformed.csv", data, "CSV row")


def test_csv_rejects_rows_wider_than_the_header() -> None:
    data = _csv_bytes([_row()]).rstrip(b"\n") + b",unexpected\n"

    _assert_import_error("wide.csv", data, "CSV row 2", "16")


@pytest.mark.parametrize("columns", [CSV_COLUMNS[:-1], CSV_COLUMNS + ("extra",)])
def test_csv_rejects_missing_or_extra_columns(columns: tuple[str, ...]) -> None:
    data = (",".join(columns) + "\n").encode()

    _assert_import_error("columns.csv", data, "CSV header", "expected exactly")


def test_csv_rejects_bad_curve_kind_with_row() -> None:
    _assert_import_error(
        "enum.csv", _csv_bytes([_row(curve_kind="unknown")]), "CSV row 2", "curve_kind"
    )


@pytest.mark.parametrize("field", ["frequency_hz", "temperature_c", "x", "y"])
def test_csv_rejects_nonfinite_numbers_with_row(field: str) -> None:
    _assert_import_error("nonfinite.csv", _csv_bytes([_row(**{field: "nan"})]), "CSV row 2", field)


def test_csv_rejects_inconsistent_repeated_metadata() -> None:
    _assert_import_error(
        "metadata.csv",
        _csv_bytes([_row(), _row(source_page=8, x=1, y=1)]),
        "CSV row 3",
        "metadata",
    )


def test_csv_rejects_inconsistent_series_metadata() -> None:
    _assert_import_error(
        "series.csv",
        _csv_bytes([_row(), _row(temperature_c=100, x=1, y=1)]),
        "CSV row 3",
        "inconsistent",
    )


def test_xlsx_rejects_missing_required_sheet() -> None:
    _assert_import_error(
        "missing.xlsx", _workbook_bytes(remove_sheet="Loss Curves"), "workbook sheets"
    )


def test_xlsx_rejects_boolean_numeric_cell() -> None:
    _assert_import_error(
        "boolean.xlsx",
        _workbook_bytes(cell_value=("B-H Curves", "D2", True)),
        "B-H Curves!A2:G2",
        "finite number",
    )


def test_xlsx_rejects_nonfinite_numeric_cell() -> None:
    _assert_import_error(
        "nonfinite.xlsx",
        _workbook_bytes(cell_value=("Loss Curves", "G2", float("inf"))),
        "Loss Curves!A2:H2",
        "finite number",
    )


def test_xlsx_rejects_empty_curve_tables() -> None:
    _assert_import_error(
        "empty.xlsx", _workbook_bytes(empty_tables=True), "B-H Curves!A2", "data row"
    )


def test_import_exported_workbook_as_draft_recomputes_fit_and_preserves_base() -> None:
    source = SourceProvenance(
        SourceKind.CSV,
        "base.csv",
        "a" * 64,
        "https://example.com/base",
        2,
        "2026-07-18T12:00:00+00:00",
        "Base source",
    )
    loss_series = tuple(
        PointSeries(
            f"loss-{frequency // 1000}khz",
            SeriesKind.LOSS_TABLE,
            "kG",
            "mW/cm3",
            CurveConditions(float(frequency), 25.0, 0.0),
            (CurvePoint(0.05, first), CurvePoint(0.1, second)),
            source.filename,
        )
        for frequency, first, second in (
            (100_000, 1000.0, 2000.0),
            (200_000, 2000.0, 4000.0),
        )
    )
    base = MaterialRecord(
        MaterialRef("Example", "Ferrite", "F1"),
        "0123456789ab",
        MaterialStatus.APPROVED,
        "2026-07-18T12:00:00+00:00",
        "reviewer@example.com",
        "approver@example.com",
        (source,),
        loss_series,
        None,
        None,
        "Approved base",
    )
    unchanged = deepcopy(base)
    exported = export_material_record_xlsx(base)
    workbook = load_workbook(io.BytesIO(exported.data))
    workbook["Loss Curves"]["G2"] = 4.0
    stream = io.BytesIO()
    workbook.save(stream)

    imported = import_material_file_as_draft(
        exported.filename,
        stream.getvalue(),
        created_at="2026-07-18T13:00:00+00:00",
        notes="Edited workbook",
    )

    assert imported.record.status is MaterialStatus.DRAFT
    assert imported.record.revision_id != base.revision_id
    assert len(imported.record.revision_id) == 12
    assert imported.record.series[0].points[0] == CurvePoint(0.0, 0.0)
    assert imported.record.series[0].points[1] == CurvePoint(0.05, 4000.0)
    assert imported.record.steinmetz is not None
    assert imported.record.notes == "Edited workbook"
    assert {name for name, _ in imported.source_files} == {
        source.filename for source in imported.record.sources
    }
    assert base == unchanged


def test_import_material_file_as_imported_persists_solver_loss_origin_in_record() -> None:
    template = material_import_template("xlsx")

    imported = import_material_file_as_imported(
        template.filename,
        template.data,
        created_at="2026-07-23T10:00:00+00:00",
    )

    assert imported.record.status is MaterialStatus.IMPORTED
    for series in imported.record.series:
        if series.kind is SeriesKind.LOSS_TABLE:
            assert series.points[0] == CurvePoint(0.0, 0.0)
