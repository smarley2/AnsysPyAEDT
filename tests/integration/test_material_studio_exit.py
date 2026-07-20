from __future__ import annotations

import json
from copy import deepcopy
from dataclasses import replace
from pathlib import Path

import pytest
from openpyxl import load_workbook
from PySide6.QtCore import QUrl

from inductor_designer.adapters.materials import FileOverlayMaterialRepository
from inductor_designer.adapters.persistence.project_repository import ProjectRepository
from inductor_designer.adapters.persistence.schema_repository import SchemaRepository
from inductor_designer.application.services.maxwell_export import (
    MaxwellExportBlocked,
    export_femm2d,
    export_maxwell3d,
    femm_manifest_json,
    generation_manifest_json,
)
from inductor_designer.domain.aedt_target import AedtEdition, AedtRelease, ModelDimension
from inductor_designer.domain.project import MaterialRevisionSelection
from inductor_designer.materials.identity import MaterialRef
from inductor_designer.materials.records import CurvePoint, SeriesKind
from inductor_designer.simulation.capabilities import (
    CapabilityReviewStatus,
    CapabilitySnapshot,
)
from inductor_designer.ui.generation_controller import CurrentProjectProvider
from inductor_designer.ui.generation_lines import GenerationBackend, run_generation
from inductor_designer.ui.main import _persist_and_publish_project
from inductor_designer.ui.material_studio_controller import MaterialStudioController
from tests.fakes.femm_solver import RecordingFemmSolver
from tests.fakes.maxwell2d_exporter import RecordingMaxwell2dExporter
from tests.fakes.maxwell_exporter import RecordingMaxwell3dExporter
from tests.unit.application.test_geometry_model import CATALOG

pytestmark = pytest.mark.ui

ROOT = Path(__file__).resolve().parents[2]
PROJECT_FIXTURE = ROOT / "tests" / "fixtures" / "sample_geometry_project.inductor.json"
CAPABILITIES = CapabilitySnapshot(
    release=AedtRelease(2025, 2),
    edition=AedtEdition.COMMERCIAL,
    include_dc_fields_3d=True,
    discovered_limits=(),
    evidence_source="recording Material Studio exit test",
    review_status=CapabilityReviewStatus.REVIEWED,
)


def _file_url(path: Path) -> str:
    return QUrl.fromLocalFile(str(path)).toString()


def _set_material_metadata(workbook_path: Path) -> None:
    workbook = load_workbook(workbook_path)
    material = workbook["Material"]
    values = {
        "manufacturer": "Magnetics",
        "material_name": "Kool Mu",
        "grade": "60",
        "source_description": "Synthetic Material Studio exit evidence",
    }
    for row in range(2, material.max_row + 1):
        field = material.cell(row, 1).value
        if field in values:
            material.cell(row, 2).value = values[field]
    workbook.save(workbook_path)


def _edit_exported_workbook(workbook_path: Path) -> None:
    workbook = load_workbook(workbook_path)
    loss = workbook["Loss Curves"]
    loss["H2"] = 110.0
    bh = workbook["B-H Curves"]
    source_rows = tuple(bh.iter_rows(min_row=2, values_only=True))
    for row in source_rows:
        copied = list(row)
        copied[0] = "bh-100c"
        copied[1] = 100.0
        bh.append(copied)
    workbook.save(workbook_path)


def test_spreadsheet_workflow_pins_exact_revision_and_series_in_recording_exports(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    fixture_document = json.loads(PROJECT_FIXTURE.read_text(encoding="utf-8"))
    assert fixture_document["schemaVersion"] == 3

    overlay = tmp_path / "overlay"
    materials = FileOverlayMaterialRepository(overlay)
    projects = ProjectRepository(SchemaRepository(ROOT / "schemas"))
    project_path = tmp_path / "material-studio.inductor.json"
    projects.save(projects.load(PROJECT_FIXTURE), project_path)
    provider = CurrentProjectProvider(projects.load(project_path))
    controller = MaterialStudioController(
        materials,
        project=provider.current(),
        project_save_callback=lambda project: _persist_and_publish_project(
            project,
            lambda value: projects.save(value, project_path),
            provider,
        ),
        now=lambda: "2026-07-19T12:00:00+00:00",
    )

    template_path = tmp_path / "material-template.xlsx"
    controller.downloadTemplate("xlsx", _file_url(template_path))
    assert template_path.read_bytes().startswith(b"PK")
    _set_material_metadata(template_path)
    controller.importTable(_file_url(template_path))
    assert controller.selectedRevision["status"] == "draft"
    controller.saveDraft()
    controller.reviewDraft("reviewer@example.com")
    controller.approveRevision("approver@example.com")

    base_revision = str(controller.selectedRevision["revisionId"])
    ref = MaterialRef("Magnetics", "Kool Mu", "60")
    base_record = deepcopy(materials.get(ref, base_revision))
    exported_path = tmp_path / "selected-revision.xlsx"
    controller.exportSelectedWorkbook(_file_url(exported_path))
    _edit_exported_workbook(exported_path)
    controller.importEditedWorkbook(_file_url(exported_path))

    edited_revision = str(controller.selectedRevision["revisionId"])
    assert controller.selectedRevision["status"] == "draft"
    assert edited_revision != base_revision
    assert materials.get(base_record.ref, base_revision) == base_record
    controller.saveDraft()
    controller.reviewDraft("reviewer-2@example.com")
    controller.approveRevision("approver-2@example.com")

    assert controller.selectedRevision["status"] == "approved"
    assert {item["revisionId"] for item in controller.revisions} == {
        base_revision,
        edited_revision,
    }
    assert sum(bool(item["isLatestApproved"]) for item in controller.revisions) == 1
    assert materials.get(base_record.ref, base_revision) == base_record
    edited = materials.get(base_record.ref, edited_revision)
    assert edited.sources[: len(base_record.sources)] == base_record.sources
    base_source_bytes = materials.source_bytes(base_record.ref, base_revision)
    edited_source_bytes = materials.source_bytes(base_record.ref, edited_revision)
    assert all(edited_source_bytes[name] == data for name, data in base_source_bytes.items())
    edited_loss = next(
        series for series in edited.series if series.series_id == "loss-100khz"
    )
    assert edited_loss.points[0] == CurvePoint(0.05, 110_000.0)
    assert {
        series.series_id
        for series in edited.series
        if series.kind is SeriesKind.BH_CURVE
    } == {"bh-25c", "bh-100c"}

    exporter = RecordingMaxwell3dExporter()
    unselected_project = replace(
        projects.load(PROJECT_FIXTURE),
        materials=(
            MaterialRevisionSelection(ref, edited_revision, edited, None),
        ),
    )
    with pytest.raises(MaxwellExportBlocked, match="multiple B-H series"):
        export_maxwell3d(
            unselected_project,
            CATALOG,
            exporter,
            tmp_path / "blocked-maxwell",
            capabilities=CAPABILITIES,
        )
    assert exporter.requests == []

    before_pin = project_path.read_bytes()
    controller.useInProject("")
    assert "multiple B-H series" in controller.statusMessage
    assert project_path.read_bytes() == before_pin

    from inductor_designer.adapters.persistence import project_repository as project_module

    real_replace = project_module.os.replace
    monkeypatch.setattr(
        project_module.os,
        "replace",
        lambda _source, _destination: (_ for _ in ()).throw(OSError("replace failed")),
    )
    controller.useInProject("bh-100c")
    assert controller.statusMessage == "replace failed"
    assert project_path.read_bytes() == before_pin
    assert provider.current().materials == ()
    monkeypatch.setattr(project_module.os, "replace", real_replace)

    controller.useInProject("bh-100c")
    persisted_document = json.loads(project_path.read_text(encoding="utf-8"))
    assert persisted_document["schemaVersion"] == 4
    assert persisted_document["materials"][0]["revisionId"] == edited_revision
    assert persisted_document["materials"][0]["bhSeriesId"] == "bh-100c"
    reloaded = projects.load(project_path)
    selection = reloaded.materials[0]
    assert selection.snapshot == edited
    assert selection.revision_id == edited_revision
    assert selection.bh_series_id == "bh-100c"

    same_session_exporter = RecordingMaxwell3dExporter()
    lines = run_generation(
        GenerationBackend.MAXWELL_3D,
        provider.current(),
        CATALOG,
        CAPABILITIES,
        tmp_path / "same-session-maxwell",
        maxwell3d_exporter=same_session_exporter,
        maxwell2d_exporter=RecordingMaxwell2dExporter(),
        femm_solver=RecordingFemmSolver(),
    )
    assert lines[-1].startswith("save: ok")
    same_session_material = same_session_exporter.requests[0].plan.core.material
    assert same_session_material.material_revision == edited_revision
    assert same_session_material.bh_series_id == "bh-100c"
    assert same_session_material.bh_curve == tuple(
        (point.y, point.x)
        for point in next(
            item for item in edited.series if item.series_id == "bh-100c"
        ).points
    )

    maxwell = export_maxwell3d(
        reloaded,
        CATALOG,
        RecordingMaxwell3dExporter(),
        tmp_path / "maxwell",
        capabilities=CAPABILITIES,
    )
    femm = export_femm2d(
        replace(reloaded, dimension_mode=ModelDimension.TWO_D),
        CATALOG,
        RecordingFemmSolver(),
        tmp_path / "femm",
        capabilities=CAPABILITIES,
        analyze=False,
    )
    for manifest in (
        json.loads(generation_manifest_json(maxwell)),
        json.loads(femm_manifest_json(femm)),
    ):
        assert manifest["coreMaterial"]["materialRevision"] == edited_revision
        assert manifest["coreMaterial"]["bhSeriesId"] == "bh-100c"
        assert manifest["coreMaterial"]["bhPointCount"] == 3


def test_latest_suggestion_and_immutable_lifecycle_edits_never_pin_implicitly(
    tmp_path: Path,
) -> None:
    repository = FileOverlayMaterialRepository(tmp_path / "overlay")
    controller = MaterialStudioController(
        repository,
        now=lambda: "2026-07-19T14:00:00+00:00",
    )
    workbook_path = tmp_path / "material.xlsx"
    controller.downloadTemplate("xlsx", _file_url(workbook_path))
    _set_material_metadata(workbook_path)
    controller.importTable(_file_url(workbook_path))
    controller.saveDraft()
    controller.reviewDraft("reviewer@example.com")

    ref = MaterialRef("Magnetics", "Kool Mu", "60")
    revision_id = str(controller.selectedRevision["revisionId"])
    reviewed = deepcopy(repository.get(ref, revision_id))
    controller.setCanonicalPoint("bh-25c", 1, 90.0, 0.09)
    assert controller.selectedRevision["status"] == "draft"
    assert controller.selectedRevision["revisionId"] != revision_id
    assert repository.get(ref, revision_id) == reviewed

    assert controller.discardChanges()
    controller.approveRevision("approver@example.com")
    approved = deepcopy(repository.get(ref, revision_id))
    controller.setCanonicalPoint("bh-25c", 1, 95.0, 0.095)
    assert controller.selectedRevision["status"] == "draft"
    assert controller.selectedRevision["revisionId"] != revision_id
    assert repository.get(ref, revision_id) == approved

    persisted_projects = []
    library = MaterialStudioController(
        repository,
        project=ProjectRepository(SchemaRepository(ROOT / "schemas")).load(
            PROJECT_FIXTURE
        ),
        project_save_callback=persisted_projects.append,
    )
    assert library.selectMaterial(ref.manufacturer, ref.name, ref.grade)
    assert library.selectedRevision == {}
    assert [item["isLatestApproved"] for item in library.revisions] == [True]
    assert persisted_projects == []
