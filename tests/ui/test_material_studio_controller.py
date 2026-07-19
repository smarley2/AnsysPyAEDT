from __future__ import annotations

import base64
import os
from collections.abc import Callable
from dataclasses import replace
from pathlib import Path

import pytest
from openpyxl import load_workbook

os.environ.setdefault("QT_QPA_PLATFORM", "offscreen")
os.environ.setdefault("QSG_RHI_BACKEND", "software")

pytest.importorskip("PySide6")

from PySide6.QtCore import QBuffer, QIODevice, QRectF, QUrl  # noqa: E402
from PySide6.QtGui import QColor, QGuiApplication, QPainter, QPdfWriter  # noqa: E402

from inductor_designer.adapters.materials import (  # noqa: E402
    export_material_record_xlsx,
    import_material_file_as_draft,
    material_import_template,
)
from inductor_designer.application.services.material_drafts import (  # noqa: E402
    ImageSeriesInput,
    MaterialDraftSession,
    approve_material_session,
    image_draft_session,
    review_material_session,
    save_material_session,
    session_from_import,
)
from inductor_designer.application.services.material_selection import (  # noqa: E402
    pin_material_revision as authoritative_pin_material_revision,
)
from inductor_designer.domain.project import InductorProject  # noqa: E402
from inductor_designer.materials.calibration import (  # noqa: E402
    AxisCalibration,
    AxisScale,
    CropRegion,
    ExtractionRecord,
    PixelPoint,
)
from inductor_designer.materials.identity import MaterialRef  # noqa: E402
from inductor_designer.materials.records import (  # noqa: E402
    CurveConditions,
    MaterialRecord,
    MaterialStatus,
    SeriesKind,
)
from inductor_designer.ui import material_studio_controller as controller_module  # noqa: E402
from inductor_designer.ui.main import create_engine  # noqa: E402
from inductor_designer.ui.material_source import render_material_source  # noqa: E402
from inductor_designer.ui.material_studio_controller import (  # noqa: E402
    MaterialStudioController,
)
from tests.fakes.material_repository import InMemoryMaterialRepository  # noqa: E402
from tests.unit.domain.test_project import make_project  # noqa: E402

_CREATED_AT = "2026-07-19T09:00:00+00:00"
_IMAGE = Path(__file__).parents[1] / "fixtures" / "materials" / "manual-bh.png"


def _file_url(path: Path) -> str:
    return QUrl.fromLocalFile(str(path)).toString()


def _session_from_upload(
    filename: str,
    data: bytes,
    *,
    created_at: str,
) -> MaterialDraftSession:
    imported = import_material_file_as_draft(filename, data, created_at=created_at)
    return session_from_import(imported.record, imported.source_files)


def _two_page_pdf() -> bytes:
    output = QBuffer()
    assert output.open(QIODevice.OpenModeFlag.WriteOnly)
    writer = QPdfWriter(output)
    writer.setResolution(72)
    painter = QPainter(writer)
    painter.fillRect(QRectF(0.0, 0.0, writer.width(), writer.height()), QColor("red"))
    assert writer.newPage()
    painter.fillRect(QRectF(0.0, 0.0, writer.width(), writer.height()), QColor("blue"))
    painter.end()
    return bytes(output.data())


def _image_session(
    ref: MaterialRef,
    filename: str,
    data: bytes,
    *,
    url: str,
    page: int | None,
    series_id: str,
) -> MaterialDraftSession:
    rendered = render_material_source(filename, data, page_index=page or 0)
    extraction = ExtractionRecord(
        CropRegion(0, 0, rendered.width_px, rendered.height_px),
        AxisCalibration(
            AxisScale.LINEAR,
            0.0,
            0.0,
            float(rendered.width_px),
            1.0,
        ),
        AxisCalibration(
            AxisScale.LINEAR,
            float(rendered.height_px),
            0.0,
            0.0,
            1.0,
        ),
        (
            PixelPoint(0.0, float(rendered.height_px)),
            PixelPoint(float(rendered.width_px), 0.0),
        ),
    )
    return image_draft_session(
        ImageSeriesInput(
            ref=ref,
            source_filename=filename,
            source_data=data,
            source_url=url,
            source_page=page,
            captured_at=_CREATED_AT,
            source_description=f"Stored {filename}",
            series_id=series_id,
            kind=SeriesKind.BH_CURVE,
            x_unit="A/m",
            y_unit="T",
            conditions=CurveConditions(None, 25.0, 0.0),
            extraction=extraction,
            created_at=_CREATED_AT,
        )
    )


def _approved_material(
    repository: InMemoryMaterialRepository,
) -> MaterialDraftSession:
    template = material_import_template("csv")
    draft = _session_from_upload(template.filename, template.data, created_at=_CREATED_AT)
    save_material_session(repository, draft)
    reviewed = review_material_session(repository, draft, "reviewer@example.com")
    return approve_material_session(repository, reviewed, "approver@example.com")


def _controller(
    repository: InMemoryMaterialRepository,
    *,
    with_project: bool = True,
) -> tuple[MaterialStudioController, list[InductorProject]]:
    saved_projects = []
    controller = MaterialStudioController(
        repository,
        project=make_project() if with_project else None,
        project_save_callback=saved_projects.append if with_project else None,
        now=lambda: "2026-07-19T10:00:00+00:00",
    )
    return controller, saved_projects


def _create_image_draft(controller: MaterialStudioController) -> None:
    controller.importSourceImage(_file_url(_IMAGE), 0)
    controller.setCrop(0, 0, 12, 8)
    controller.setXAxis("linear", 1.0, 0.0, 11.0, 2.0)
    controller.setYAxis("linear", 7.0, 0.0, 1.0, 2.0)
    controller.setSeriesMetadata(
        "bh-manual",
        "bh-curve",
        "Oe",
        "kG",
        float("nan"),
        0.0,
        float("nan"),
    )
    controller.addPixelPoint(1.0, 7.0)
    controller.addPixelPoint(11.0, 1.0)
    controller.createImageDraft(
        "Example",
        "Ferrite",
        "Manual",
        "Synthetic manual B-H curve",
    )


@pytest.mark.ui
def test_library_refresh_and_latest_approved_are_advisory_only() -> None:
    repository = InMemoryMaterialRepository()
    approved = _approved_material(repository)

    controller, _ = _controller(repository)

    assert controller.materials == [
        {
            "manufacturer": approved.record.ref.manufacturer,
            "name": approved.record.ref.name,
            "grade": approved.record.ref.grade,
        }
    ]
    assert controller.revisions == []
    assert controller.selectedRevision == {}

    controller.selectMaterial(
        approved.record.ref.manufacturer,
        approved.record.ref.name,
        approved.record.ref.grade,
    )

    assert [item["revisionId"] for item in controller.revisions] == [
        approved.record.revision_id
    ]
    assert controller.revisions[0]["isLatestApproved"] is True
    assert controller.selectedRevision == {}
    assert controller.dirty is False


@pytest.mark.ui
def test_selected_revision_exposes_plain_details_and_action_flags() -> None:
    repository = InMemoryMaterialRepository()
    approved = _approved_material(repository)
    controller, _ = _controller(repository)
    controller.selectMaterial(
        approved.record.ref.manufacturer,
        approved.record.ref.name,
        approved.record.ref.grade,
    )

    controller.selectRevision(approved.record.revision_id)

    assert controller.selectedRevision["revisionId"] == approved.record.revision_id
    assert controller.selectedRevision["status"] == "approved"
    assert controller.selectedRevision["reviewedBy"] == "reviewer@example.com"
    assert controller.selectedRevision["approvedBy"] == "approver@example.com"
    assert all(isinstance(item, dict) for item in controller.series)
    assert all(isinstance(item, dict) for item in controller.points)
    assert all(isinstance(item, dict) for item in controller.issues)
    assert isinstance(controller.fit, dict)
    assert controller.canSave is False
    assert controller.canReview is False
    assert controller.canApprove is False
    assert controller.canUseInProject is True


@pytest.mark.ui
def test_import_save_review_and_approve_refresh_state(tmp_path: Path) -> None:
    repository = InMemoryMaterialRepository()
    approved = _approved_material(repository)
    controller, _ = _controller(repository)
    template = material_import_template("csv")
    edited_bytes = template.data.replace(
        b"Synthetic example data - replace before review",
        b"Controller lifecycle draft",
    )
    source = tmp_path / template.filename
    source.write_bytes(edited_bytes)

    controller.importTable(_file_url(source))

    draft_id = controller.selectedRevision["revisionId"]
    assert draft_id != approved.record.revision_id
    assert controller.selectedRevision["status"] == "draft"
    assert controller.dirty is True
    assert controller.canSave is True
    assert controller.canReview is False
    assert controller.canApprove is False

    controller.saveDraft()

    assert controller.dirty is False
    assert controller.canSave is False
    assert controller.canReview is True
    assert {item["revisionId"] for item in controller.revisions} == {
        approved.record.revision_id,
        draft_id,
    }

    selected_before = dict(controller.selectedRevision)
    revisions_before = list(controller.revisions)
    controller.reviewDraft("  ")
    assert controller.selectedRevision == selected_before
    assert controller.revisions == revisions_before
    assert controller.statusMessage == "Reviewer must not be blank."

    controller.reviewDraft("reviewer-2@example.com")

    assert controller.selectedRevision["status"] == "reviewed"
    assert controller.selectedRevision["reviewedBy"] == "reviewer-2@example.com"
    assert controller.canReview is False
    assert controller.canApprove is True

    selected_before = dict(controller.selectedRevision)
    controller.approveRevision("")
    assert controller.selectedRevision == selected_before
    assert controller.statusMessage == "Approver must not be blank."

    controller.approveRevision("approver-2@example.com")

    assert controller.selectedRevision["status"] == "approved"
    assert controller.selectedRevision["approvedBy"] == "approver-2@example.com"
    assert controller.canApprove is False
    assert controller.canUseInProject is True
    approved_summary = next(
        item for item in controller.revisions if item["revisionId"] == draft_id
    )
    assert approved_summary["isLatestApproved"] is True


@pytest.mark.ui
def test_known_selection_error_keeps_prior_selection() -> None:
    repository = InMemoryMaterialRepository()
    approved = _approved_material(repository)
    controller, _ = _controller(repository)
    controller.selectMaterial(
        approved.record.ref.manufacturer,
        approved.record.ref.name,
        approved.record.ref.grade,
    )
    controller.selectRevision(approved.record.revision_id)
    before = dict(controller.selectedRevision)

    controller.selectRevision("missing")

    assert controller.selectedRevision == before
    assert "unknown material revision:" in controller.statusMessage


@pytest.mark.ui
def test_known_errors_are_logged_with_traceback(
    caplog: pytest.LogCaptureFixture,
) -> None:
    repository = InMemoryMaterialRepository()
    approved = _approved_material(repository)
    controller, _ = _controller(repository)
    controller.selectMaterial(
        approved.record.ref.manufacturer,
        approved.record.ref.name,
        approved.record.ref.grade,
    )
    caplog.set_level("ERROR", logger="inductor_designer.ui.material_studio_controller")

    controller.selectRevision("missing")

    assert "unknown material revision:" in controller.statusMessage
    assert any(
        record.exc_info is not None and "Material Studio action failed" in record.message
        for record in caplog.records
    )


@pytest.mark.ui
def test_qml_property_values_are_deep_copies() -> None:
    repository = InMemoryMaterialRepository()
    approved = _approved_material(repository)
    controller, _ = _controller(repository)
    controller.selectMaterial(
        approved.record.ref.manufacturer,
        approved.record.ref.name,
        approved.record.ref.grade,
    )
    controller.selectRevision(approved.record.revision_id)

    materials = controller.materials
    revisions = controller.revisions
    selected = controller.selectedRevision
    materials[0]["name"] = "mutated"
    revisions[0]["status"] = "mutated"
    selected["sources"][0]["filename"] = "mutated"

    assert controller.materials[0]["name"] == approved.record.ref.name
    assert controller.revisions[0]["status"] == "approved"
    assert controller.selectedRevision["sources"][0]["filename"] != "mutated"


@pytest.mark.ui
def test_downloads_require_local_destinations_and_use_exact_bytes(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    repository = InMemoryMaterialRepository()
    approved = _approved_material(repository)
    controller, _ = _controller(repository)

    controller.downloadTemplate("csv", "")
    assert list(tmp_path.iterdir()) == []

    template_target = tmp_path / "template.csv"
    controller.downloadTemplate("csv", _file_url(template_target))
    assert template_target.read_bytes() == material_import_template("csv").data

    controller.selectMaterial(
        approved.record.ref.manufacturer,
        approved.record.ref.name,
        approved.record.ref.grade,
    )
    controller.selectRevision(approved.record.revision_id)
    workbook_target = tmp_path / "selected.xlsx"
    expected_workbook = export_material_record_xlsx(
        approved.record,
        exported_at="2026-07-19T10:00:00+00:00",
    )
    monkeypatch.setattr(
        controller_module,
        "export_material_record_xlsx",
        lambda _record, *, exported_at: expected_workbook,
    )

    controller.exportSelectedWorkbook("")
    assert not workbook_target.exists()
    controller.exportSelectedWorkbook(_file_url(workbook_target))

    assert workbook_target.read_bytes() == expected_workbook.data


@pytest.mark.ui
def test_table_and_workbook_uploads_read_before_replacing_state(tmp_path: Path) -> None:
    repository = InMemoryMaterialRepository()
    approved = _approved_material(repository)
    controller, _ = _controller(repository)
    controller.selectMaterial(
        approved.record.ref.manufacturer,
        approved.record.ref.name,
        approved.record.ref.grade,
    )
    controller.selectRevision(approved.record.revision_id)
    selected_before = dict(controller.selectedRevision)

    controller.importTable("")
    controller.importTable("https://example.com/material.csv")
    controller.importEditedWorkbook(_file_url(tmp_path / "missing.xlsx"))

    assert controller.selectedRevision == selected_before
    assert controller.dirty is False

    workbook = tmp_path / "edited.xlsx"
    workbook.write_bytes(export_material_record_xlsx(approved.record).data)
    controller.importEditedWorkbook(_file_url(workbook))

    assert controller.selectedRevision["status"] == "draft"
    assert controller.selectedRevision["revisionId"] != approved.record.revision_id
    assert controller.dirty is True
    assert controller.canSave is True


@pytest.mark.ui
def test_edited_workbook_requires_matching_clean_selected_base(tmp_path: Path) -> None:
    repository = InMemoryMaterialRepository()
    approved = _approved_material(repository)
    controller, _ = _controller(repository)
    controller.selectMaterial(
        approved.record.ref.manufacturer,
        approved.record.ref.name,
        approved.record.ref.grade,
    )
    controller.selectRevision(approved.record.revision_id)
    selected_before = dict(controller.selectedRevision)

    wrong = tmp_path / "wrong-base.xlsx"
    wrong.write_bytes(export_material_record_xlsx(approved.record).data)
    workbook = load_workbook(wrong)
    lineage = workbook["_MaterialStudio"]
    lineage["B7"] = "000000000000"
    workbook.save(wrong)

    controller.importEditedWorkbook(_file_url(wrong))

    assert "does not match the selected revision" in controller.statusMessage
    assert controller.selectedRevision == selected_before
    assert controller.dirty is False

    generic = tmp_path / "generic.xlsx"
    generic.write_bytes(material_import_template("xlsx").data)
    controller.importEditedWorkbook(_file_url(generic))
    assert "does not identify an exported base revision" in controller.statusMessage
    assert controller.selectedRevision == selected_before
    assert controller.dirty is False


@pytest.mark.ui
def test_image_import_exposes_rendered_png_and_preserves_selection_on_error() -> None:
    repository = InMemoryMaterialRepository()
    approved = _approved_material(repository)
    controller, _ = _controller(repository)
    controller.selectMaterial(
        approved.record.ref.manufacturer,
        approved.record.ref.name,
        approved.record.ref.grade,
    )
    controller.selectRevision(approved.record.revision_id)
    source_bytes = _IMAGE.read_bytes()
    rendered = render_material_source(_IMAGE.name, source_bytes)

    controller.importSourceImage(_file_url(_IMAGE), 0)

    assert controller.source == {
        "dataUrl": "data:image/png;base64,"
        + base64.b64encode(rendered.png_data).decode("ascii"),
        "filename": _IMAGE.name,
        "width": rendered.width_px,
        "height": rendered.height_px,
        "pageCount": 1,
        "pageIndex": 0,
    }
    assert controller.selectedRevision == {}
    assert controller.series == []
    assert controller.canSave is False
    assert controller.canReview is False
    assert controller.canApprove is False
    assert controller.canUseInProject is False
    assert controller.dirty is True
    source_before = dict(controller.source)
    selected_before = dict(controller.selectedRevision)

    controller.importSourceImage("", 0)
    controller.importSourceImage("https://example.com/manual-bh.png", 0)

    assert controller.source == source_before
    assert controller.selectedRevision == selected_before


@pytest.mark.ui
def test_loaded_source_blocks_stale_lifecycle_until_matching_draft_is_created() -> None:
    repository = InMemoryMaterialRepository()
    approved = _approved_material(repository)
    controller, saved_projects = _controller(repository)
    controller.selectMaterial(
        approved.record.ref.manufacturer,
        approved.record.ref.name,
        approved.record.ref.grade,
    )
    controller.selectRevision(approved.record.revision_id)
    assert controller.canUseInProject is True

    source_bytes = _IMAGE.read_bytes()
    controller.importSourceImage(_file_url(_IMAGE), 0)
    controller.saveDraft()
    controller.reviewDraft("reviewer-2@example.com")
    controller.approveRevision("approver-2@example.com")
    controller.useInProject("bh-25c")

    assert saved_projects == []
    assert repository.list_revisions(approved.record.ref) == (
        approved.record.revision_id,
    )
    assert controller.selectedRevision == {}
    assert controller.canSave is False
    assert controller.canReview is False
    assert controller.canApprove is False
    assert controller.canUseInProject is False

    controller.setXAxis("linear", 1.0, 0.0, 11.0, 2.0)
    controller.setYAxis("linear", 7.0, 0.0, 1.0, 2.0)
    controller.setSeriesMetadata(
        "bh-new-source",
        "bh-curve",
        "A/m",
        "T",
        float("nan"),
        25.0,
        0.0,
    )
    controller.addPixelPoint(1.0, 7.0)
    controller.addPixelPoint(11.0, 1.0)
    controller.createImageDraft(
        "Replacement",
        "Source",
        "R1",
        "Replacement image",
    )
    assert controller.canSave is True

    controller.saveDraft()

    ref = MaterialRef("Replacement", "Source", "R1")
    revision_id = str(controller.selectedRevision["revisionId"])
    assert repository.source_bytes(ref, revision_id) == {_IMAGE.name: source_bytes}


@pytest.mark.ui
@pytest.mark.parametrize(
    ("filename", "data_factory", "page"),
    [
        ("stored.png", lambda: _IMAGE.read_bytes(), None),
        ("stored.pdf", _two_page_pdf, 1),
    ],
)
def test_select_revision_restores_exact_stored_image_source_and_editor_snapshot(
    filename: str,
    data_factory: Callable[[], bytes],
    page: int | None,
) -> None:
    repository = InMemoryMaterialRepository()
    data = data_factory()
    ref = MaterialRef("Stored", "Image", filename)
    url = f"https://example.com/{filename}"
    session = _image_session(
        ref,
        filename,
        data,
        url=url,
        page=page,
        series_id="bh-stored",
    )
    save_material_session(repository, session)
    rendered = render_material_source(filename, data, page_index=page or 0)
    controller, _ = _controller(repository)
    controller.selectMaterial(ref.manufacturer, ref.name, ref.grade)

    controller.selectRevision(session.record.revision_id)

    assert controller.source["dataUrl"] == (
        "data:image/png;base64,"
        + base64.b64encode(rendered.png_data).decode("ascii")
    )
    assert controller.source["width"] == rendered.width_px
    assert controller.source["height"] == rendered.height_px
    assert controller.source["pageIndex"] == (page or 0)
    assert controller.source["url"] == url
    assert controller.source["page"] == page
    assert controller.source["capturedAt"] == _CREATED_AT
    assert controller.source["description"] == f"Stored {filename}"
    assert controller.source["sha256"] == session.record.sources[0].sha256
    assert controller.sourceComparisonAvailable is True
    assert controller.sourcePoints == controller.points
    assert controller.imageEditing["crop"] == {
        "left": 0,
        "top": 0,
        "width": rendered.width_px,
        "height": rendered.height_px,
    }

    source_snapshot = controller.source
    extraction_snapshot = controller.imageEditing
    controller.setCrop(1, 1, rendered.width_px - 1, rendered.height_px - 1)
    assert controller.discardChanges() is True
    assert controller.source == source_snapshot
    assert controller.imageEditing == extraction_snapshot


@pytest.mark.ui
def test_series_switch_restores_matching_source_and_fails_atomically(
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    repository = InMemoryMaterialRepository()
    ref = MaterialRef("Stored", "Mixed", "Sources")
    image = _image_session(
        ref,
        "stored.png",
        _IMAGE.read_bytes(),
        url="https://example.com/stored.png",
        page=None,
        series_id="bh-image",
    )
    table_template = material_import_template("csv")
    table = _session_from_upload(
        table_template.filename,
        table_template.data,
        created_at=_CREATED_AT,
    )
    table_series = replace(
        table.record.series[0],
        series_id="bh-table",
    )
    mixed_record = replace(
        image.record,
        revision_id="dddddddddddd",
        sources=(*image.record.sources, *table.record.sources),
        series=(*image.record.series, table_series),
    )
    repository.save(
        mixed_record,
        dict((*image.source_files, *table.source_files)),
    )
    controller, _ = _controller(repository)
    controller.selectMaterial(ref.manufacturer, ref.name, ref.grade)
    controller.selectRevision(mixed_record.revision_id)
    image_source = controller.source

    controller.selectSeries("bh-table")

    assert controller.source == {}
    assert controller.imageEditing["metadata"]["seriesId"] == "bh-table"
    table_snapshot = controller.imageEditing

    monkeypatch.setattr(
        controller_module,
        "render_material_source",
        lambda *_args, **_kwargs: (_ for _ in ()).throw(ValueError("render failed")),
    )
    controller.selectSeries("bh-image")

    assert controller.source == {}
    assert controller.imageEditing == table_snapshot
    assert controller.statusMessage == "render failed"

    monkeypatch.undo()
    controller.selectSeries("bh-image")
    assert controller.source == image_source
    assert controller.imageEditing["metadata"]["seriesId"] == "bh-image"


@pytest.mark.ui
def test_new_record_clears_source_while_lifecycle_preserves_it(tmp_path: Path) -> None:
    repository = InMemoryMaterialRepository()
    approved = _approved_material(repository)
    controller, _ = _controller(repository)
    controller.selectMaterial(
        approved.record.ref.manufacturer,
        approved.record.ref.name,
        approved.record.ref.grade,
    )
    controller.selectRevision(approved.record.revision_id)
    controller.importSourceImage(_file_url(_IMAGE), 0)
    assert controller.source
    source_changes: list[None] = []
    controller.sourceChanged.connect(lambda: source_changes.append(None))
    template = material_import_template("csv")
    edited = tmp_path / template.filename
    edited.write_bytes(
        template.data.replace(
            b"Synthetic example data - replace before review",
            b"Source clearing lifecycle draft",
        )
    )

    controller.importTable(_file_url(edited))

    assert controller.source
    assert "Save or discard" in controller.statusMessage
    controller.discardChanges()
    source_changes.clear()
    controller.importTable(_file_url(edited))

    assert controller.source == {}
    assert source_changes == [None]

    controller.importSourceImage(_file_url(_IMAGE), 0)
    rendered_source = controller.source
    controller.saveDraft()
    controller.reviewDraft("reviewer-2@example.com")

    assert controller.source == rendered_source


@pytest.mark.ui
def test_image_edit_state_builds_draft_and_invalid_axis_keeps_last_valid_session() -> None:
    controller, _ = _controller(InMemoryMaterialRepository())

    _create_image_draft(controller)

    assert controller.selectedRevision["manufacturer"] == "Example"
    assert controller.selectedRevision["createdAt"] == "2026-07-19T10:00:00+00:00"
    assert controller.series == [
        {
            "seriesId": "bh-manual",
            "kind": "bh-curve",
            "xUnit": "Oe",
            "yUnit": "kG",
            "frequencyHz": None,
            "temperatureC": 0.0,
            "dcBiasAPerM": None,
            "pointCount": 2,
            "imageBacked": True,
        }
    ]
    assert controller.points == [
        {"seriesId": "bh-manual", "index": 0, "x": 0.0, "y": 0.0},
        {
            "seriesId": "bh-manual",
            "index": 1,
            "x": pytest.approx(159.154943092),
            "y": 0.2,
        },
    ]
    assert controller.imageEditing["pixelPoints"] == [
        {"xPx": 1.0, "yPx": 7.0},
        {"xPx": 11.0, "yPx": 1.0},
    ]
    assert controller.canSave is True
    last_valid_revision = controller.selectedRevision["revisionId"]

    controller.setXAxis("log", 1.0, 0.0, 11.0, 2.0)
    controller.movePixelPoint(1, 10.0, 2.0)

    assert controller.imageEditing["xAxis"]["scale"] == "log"
    assert controller.imageEditing["xAxis"]["valueA"] == 0.0
    assert controller.imageEditing["pixelPoints"][1] == {"xPx": 11.0, "yPx": 1.0}
    assert controller.selectedRevision["revisionId"] == last_valid_revision
    assert controller.statusMessage == "Apply or correct the visible editor input first."
    assert controller.canSave is False
    controller.saveDraft()
    assert controller.dirty is True
    assert controller.selectedRevision["revisionId"] == last_valid_revision
    assert controller.statusMessage == "Resolve the invalid editor input before saving."

    controller.setXAxis("linear", 1.0, 0.0, 11.0, 2.0)

    assert controller.selectedRevision["revisionId"] == last_valid_revision
    assert controller.points[1]["x"] == pytest.approx(159.154943092)
    assert controller.points[1]["y"] == pytest.approx(0.2)
    assert controller.statusMessage == "Image extraction updated."
    assert controller.canSave is True


@pytest.mark.ui
def test_controller_rejects_crop_outside_loaded_source_bounds() -> None:
    controller, _ = _controller(InMemoryMaterialRepository())
    _create_image_draft(controller)
    valid_revision = controller.selectedRevision["revisionId"]

    for crop in ((-1, 0, 12, 8), (0, -1, 12, 8), (0, 0, 13, 8), (0, 0, 12, 9)):
        controller.setCrop(*crop)
        assert controller.imageEditing["crop"] == dict(
            zip(("left", "top", "width", "height"), crop, strict=True)
        )
        assert controller.selectedRevision["revisionId"] == valid_revision
        assert controller.canSave is False
        assert "source bounds" in controller.statusMessage
        controller.setCrop(0, 0, 12, 8)
        assert controller.canSave is True

    controller.setCrop(11, 7, 1, 1)
    assert controller.imageEditing["crop"] == {
        "left": 11,
        "top": 7,
        "width": 1,
        "height": 1,
    }
    assert controller.canSave is True


@pytest.mark.ui
def test_image_metadata_and_canonical_point_edits_use_authoritative_replacements() -> None:
    controller, _ = _controller(InMemoryMaterialRepository())
    _create_image_draft(controller)
    source_points = controller.sourcePoints

    controller.setSeriesMetadata(
        "bh room",
        "bh-curve",
        "A/m",
        "T",
        0.0,
        30.0,
        0.0,
    )

    assert controller.series[0]["seriesId"] == "bh room"
    assert controller.series[0]["imageBacked"] is True
    assert controller.series[0]["frequencyHz"] == 0.0
    assert controller.series[0]["temperatureC"] == 30.0
    assert controller.series[0]["dcBiasAPerM"] == 0.0
    assert controller.imageEditing["metadata"]["seriesId"] == "bh room"
    assert controller.sourcePoints == [
        {**point, "seriesId": "bh room"} for point in source_points
    ]

    controller.setCanonicalPoint("bh room", 1, 200.0, 0.3)

    assert controller.series[0]["imageBacked"] is False
    assert controller.sourcePoints != controller.points
    assert controller.points[1] == {
        "seriesId": "bh room",
        "index": 1,
        "x": 200.0,
        "y": 0.3,
    }
    assert [source["kind"] for source in controller.selectedRevision["sources"]] == [
        "image",
        "csv",
    ]
    assert controller.statusMessage == (
        "Numeric editing converted the image-backed series to a direct table edit; "
        "the original image/PDF remains as supplemental provenance."
    )


@pytest.mark.ui
def test_table_metadata_edit_treats_only_nan_as_an_absent_condition(tmp_path: Path) -> None:
    controller, _ = _controller(InMemoryMaterialRepository())
    template = material_import_template("csv")
    source = tmp_path / template.filename
    source.write_bytes(template.data)
    controller.importTable(_file_url(source))

    controller.setSeriesMetadata(
        "bh-25c",
        "bh-curve",
        "A/m",
        "T",
        float("nan"),
        0.0,
        0.0,
    )

    bh = next(item for item in controller.series if item["seriesId"] == "bh-25c")
    assert bh["frequencyHz"] is None
    assert bh["temperatureC"] == 0.0
    assert bh["dcBiasAPerM"] == 0.0
    assert bh["imageBacked"] is False
    assert controller.dirty is True


@pytest.mark.ui
@pytest.mark.parametrize(
    ("series_id", "x_unit", "message"),
    [
        ("", "A/m", "must not be blank"),
        ("loss-100khz", "A/m", "already exists"),
        ("bh-renamed", "s", "unit"),
    ],
)
def test_invalid_visible_metadata_blocks_save_without_replacing_last_valid_session(
    tmp_path: Path,
    series_id: str,
    x_unit: str,
    message: str,
) -> None:
    repository = InMemoryMaterialRepository()
    controller, _ = _controller(repository)
    template = material_import_template("csv")
    source = tmp_path / template.filename
    source.write_bytes(template.data)
    controller.importTable(_file_url(source))
    controller.setXAxis("linear", 0.0, 0.0, 1.0, 1.0)
    controller.setYAxis("linear", 1.0, 0.0, 0.0, 1.0)
    last_valid_revision = controller.selectedRevision["revisionId"]

    controller.setSeriesMetadata(
        series_id,
        "bh-curve",
        x_unit,
        "T",
        float("nan"),
        25.0,
        float("nan"),
    )

    assert controller.imageEditing["metadata"]["seriesId"] == series_id
    assert controller.imageEditing["metadata"]["xUnit"] == x_unit
    assert controller.selectedRevision["revisionId"] == last_valid_revision
    assert message in controller.statusMessage
    assert controller.dirty is True
    assert controller.canSave is False

    controller.saveDraft()

    assert repository.list_materials() == ()
    assert controller.imageEditing["metadata"]["seriesId"] == series_id
    assert controller.imageEditing["metadata"]["xUnit"] == x_unit
    assert controller.dirty is True
    assert controller.canSave is False

    controller.setCrop(0, 0, 1, 1)

    assert controller.imageEditing["metadata"]["seriesId"] == series_id
    assert controller.imageEditing["metadata"]["xUnit"] == x_unit
    assert controller.selectedRevision["revisionId"] == last_valid_revision
    assert controller.canSave is False

    controller.setSeriesMetadata(
        "bh-corrected",
        "bh-curve",
        "A/m",
        "T",
        float("nan"),
        25.0,
        float("nan"),
    )

    assert controller.series[0]["seriesId"] == "bh-corrected"
    assert controller.canSave is True


@pytest.mark.ui
def test_pending_visible_editor_input_blocks_stale_save(tmp_path: Path) -> None:
    repository = InMemoryMaterialRepository()
    controller, _ = _controller(repository)
    template = material_import_template("csv")
    source = tmp_path / template.filename
    source.write_bytes(template.data)
    controller.importTable(_file_url(source))

    controller.invalidateEditorInput(
        "metadata",
        "Apply or correct the visible editor input.",
    )
    controller.saveDraft()

    assert repository.list_materials() == ()
    assert controller.dirty is True
    assert controller.canSave is False
    assert controller.statusMessage == "Resolve the invalid editor input before saving."


@pytest.mark.ui
def test_direct_library_selection_refuses_to_replace_dirty_state(tmp_path: Path) -> None:
    repository = InMemoryMaterialRepository()
    first = _approved_material(repository)
    second_record = replace(
        first.record,
        ref=MaterialRef("Other", "Ferrite", "N97"),
        revision_id="eeeeeeeeeeee",
    )
    repository.save(second_record, dict(first.source_files))
    controller, _ = _controller(repository)
    controller.selectMaterial(
        first.record.ref.manufacturer,
        first.record.ref.name,
        first.record.ref.grade,
    )
    controller.selectRevision(first.record.revision_id)
    controller.importSourceImage(_file_url(_IMAGE), 0)
    selected_before = controller.selectedRevision
    source_before = controller.source

    controller.selectRevision("missing")
    controller.selectMaterial("Other", "Ferrite", "N97")

    assert controller.selectedRevision == selected_before
    assert controller.source == source_before
    assert controller.dirty is True
    assert controller.statusMessage == (
        "Save or discard unsaved material changes before changing library selection."
    )


@pytest.mark.ui
def test_table_point_delete_edit_uses_authoritative_table_replacement(
    tmp_path: Path,
) -> None:
    controller, _ = _controller(InMemoryMaterialRepository())
    template = material_import_template("csv")
    source = tmp_path / template.filename
    source.write_bytes(template.data)
    controller.importTable(_file_url(source))
    before = [item for item in controller.points if item["seriesId"] == "bh-25c"]

    controller.deletePoint(1)

    after = [item for item in controller.points if item["seriesId"] == "bh-25c"]
    assert len(after) == len(before) - 1
    assert controller.statusMessage == "Canonical point deleted."
    assert controller.dirty is True


@pytest.mark.ui
def test_series_selector_changes_the_active_editor_without_marking_dirty(
    tmp_path: Path,
) -> None:
    controller, _ = _controller(InMemoryMaterialRepository())
    template = material_import_template("csv")
    source = tmp_path / template.filename
    source.write_bytes(template.data)
    controller.importTable(_file_url(source))
    controller.saveDraft()

    controller.selectSeries("loss-100khz")

    assert controller.imageEditing["metadata"]["seriesId"] == "loss-100khz"
    assert controller.imageEditing["metadata"]["kind"] == "loss-table"
    assert controller.dirty is False

    before = controller.imageEditing
    controller.selectSeries("missing")
    assert controller.imageEditing == before
    assert controller.statusMessage == "Series 'missing' does not exist."


@pytest.mark.ui
def test_discard_edit_restores_last_saved_session_and_loaded_source() -> None:
    repository = InMemoryMaterialRepository()
    approved = _approved_material(repository)
    controller, _ = _controller(repository)
    controller.selectMaterial(
        approved.record.ref.manufacturer,
        approved.record.ref.name,
        approved.record.ref.grade,
    )
    controller.selectRevision(approved.record.revision_id)
    selected_before = controller.selectedRevision

    controller.importSourceImage(_file_url(_IMAGE), 0)

    assert controller.dirty is True
    assert controller.source
    assert controller.discardChanges() is True
    assert controller.dirty is False
    assert controller.source == {}
    assert controller.selectedRevision == selected_before
    assert controller.statusMessage == "Unsaved changes discarded."

    _create_image_draft(controller)
    controller.saveDraft()
    saved_revision = controller.selectedRevision["revisionId"]
    saved_points = controller.points
    controller.setCanonicalPoint("bh-manual", 1, 200.0, 0.3)

    assert controller.dirty is True
    assert controller.discardChanges() is True
    assert controller.dirty is False
    assert controller.selectedRevision["revisionId"] == saved_revision
    assert controller.points == saved_points


class _PostPersistenceListingFailureRepository(InMemoryMaterialRepository):
    fail_listing = False

    def list_materials(self) -> tuple[MaterialRef, ...]:
        if self.fail_listing:
            raise ValueError("post-persistence listing unavailable")
        return super().list_materials()


@pytest.mark.ui
def test_lifecycle_success_is_committed_before_refresh_warning(tmp_path: Path) -> None:
    repository = _PostPersistenceListingFailureRepository()
    _approved_material(repository)
    controller, _ = _controller(repository)
    template = material_import_template("csv")
    edited = tmp_path / template.filename
    edited.write_bytes(
        template.data.replace(
            b"Synthetic example data - replace before review",
            b"Post-persistence refresh draft",
        )
    )
    controller.importTable(_file_url(edited))
    controller.saveDraft()
    ref = MaterialRef(
        controller.selectedRevision["manufacturer"],
        controller.selectedRevision["name"],
        controller.selectedRevision["grade"],
    )
    revision_id = controller.selectedRevision["revisionId"]
    repository.fail_listing = True

    controller.reviewDraft("reviewer-2@example.com")

    assert repository.get(ref, revision_id).status is MaterialStatus.REVIEWED
    assert controller.selectedRevision["status"] == "reviewed"
    assert controller.selectedRevision["reviewedBy"] == "reviewer-2@example.com"
    assert controller.canApprove is True
    assert "reviewed" in controller.statusMessage
    assert "refresh failed" in controller.statusMessage


@pytest.mark.ui
def test_project_use_requires_explicit_multi_bh_id_and_saves_once() -> None:
    repository = InMemoryMaterialRepository()
    approved = _approved_material(repository)
    bh = next(item for item in approved.record.series if item.kind is SeriesKind.BH_CURVE)
    multi = replace(
        approved.record,
        revision_id="ffffffffffff",
        series=(*approved.record.series, replace(bh, series_id="bh-100c")),
    )
    repository.save(multi, dict(approved.source_files))
    controller, saved_projects = _controller(repository)
    controller.selectMaterial(
        multi.ref.manufacturer,
        multi.ref.name,
        multi.ref.grade,
    )
    controller.selectRevision(multi.revision_id)

    controller.useInProject("")

    assert saved_projects == []
    assert "multiple B-H series" in controller.statusMessage

    controller.useInProject("bh-100c")

    assert len(saved_projects) == 1
    assert saved_projects[0].materials[0].revision_id == multi.revision_id
    assert saved_projects[0].materials[0].bh_series_id == "bh-100c"


@pytest.mark.ui
@pytest.mark.parametrize(
    ("requested_id", "expected_id"),
    [("   ", None), ("  bh-25c  ", "bh-25c")],
)
def test_bh_selection_is_normalized_then_forwarded_to_authoritative_service(
    monkeypatch: pytest.MonkeyPatch,
    requested_id: str,
    expected_id: str | None,
) -> None:
    repository = InMemoryMaterialRepository()
    approved = _approved_material(repository)
    controller, saved_projects = _controller(repository)
    controller.selectMaterial(
        approved.record.ref.manufacturer,
        approved.record.ref.name,
        approved.record.ref.grade,
    )
    controller.selectRevision(approved.record.revision_id)
    forwarded: list[str | None] = []

    def recording_pin(
        project: InductorProject,
        record: MaterialRecord,
        *,
        bh_series_id: str | None,
    ) -> InductorProject:
        forwarded.append(bh_series_id)
        return authoritative_pin_material_revision(
            project,
            record,
            bh_series_id=bh_series_id,
        )

    monkeypatch.setattr(controller_module, "pin_material_revision", recording_pin)

    controller.useInProject(requested_id)

    assert forwarded == [expected_id]
    assert len(saved_projects) == 1
    assert saved_projects[0].materials[0].bh_series_id == expected_id


@pytest.mark.ui
def test_library_controller_without_project_cannot_pin() -> None:
    repository = InMemoryMaterialRepository()
    approved = _approved_material(repository)
    controller, saved_projects = _controller(repository, with_project=False)
    controller.selectMaterial(
        approved.record.ref.manufacturer,
        approved.record.ref.name,
        approved.record.ref.grade,
    )
    controller.selectRevision(approved.record.revision_id)

    assert controller.canUseInProject is False
    controller.useInProject("")
    assert saved_projects == []
    assert controller.statusMessage == "No project is loaded."


@pytest.mark.ui
def test_create_engine_injects_material_studio_controller() -> None:
    app = QGuiApplication.instance() or QGuiApplication([])
    controller, _ = _controller(InMemoryMaterialRepository(), with_project=False)

    engine = create_engine(material_studio_controller=controller)

    assert app is not None
    assert engine.rootContext().contextProperty("materialStudioController") is controller
    assert len(engine.rootObjects()) == 1


@pytest.mark.ui
def test_destructive_imports_refuse_to_replace_dirty_state(tmp_path: Path) -> None:
    controller, _ = _controller(InMemoryMaterialRepository(), with_project=False)
    template = material_import_template("csv")
    first = tmp_path / "first.csv"
    second = tmp_path / "second.csv"
    first.write_bytes(template.data)
    second.write_bytes(template.data.replace(b"Example Vendor", b"Second Vendor"))
    controller.importTable(_file_url(first))
    snapshot = (
        controller.selectedRevision,
        controller.series,
        controller.source,
    )

    controller.importTable(_file_url(second))
    controller.importSourceImage(_file_url(_IMAGE), 0)

    assert (controller.selectedRevision, controller.series, controller.source) == snapshot
    assert controller.dirty is True
    assert "Save or discard" in controller.statusMessage


@pytest.mark.ui
@pytest.mark.parametrize("download_kind", ["template", "revision"])
def test_download_failure_is_atomic_and_cleans_same_directory_temp(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
    download_kind: str,
) -> None:
    repository = InMemoryMaterialRepository()
    approved = _approved_material(repository)
    controller, _ = _controller(repository, with_project=False)
    if download_kind == "revision":
        controller.selectMaterial(
            approved.record.ref.manufacturer,
            approved.record.ref.name,
            approved.record.ref.grade,
        )
        controller.selectRevision(approved.record.revision_id)
    destination = tmp_path / ("selected.xlsx" if download_kind == "revision" else "template.csv")
    original = b"existing destination"
    destination.write_bytes(original)

    def fail_replace(_source: object, _destination: object) -> None:
        raise OSError("controlled replace failure")

    monkeypatch.setattr(controller_module.os, "replace", fail_replace)
    if download_kind == "revision":
        controller.exportSelectedWorkbook(_file_url(destination))
    else:
        controller.downloadTemplate("csv", _file_url(destination))

    assert destination.read_bytes() == original
    assert sorted(path.name for path in tmp_path.iterdir()) == [destination.name]
    assert controller.statusMessage == "controlled replace failure"


@pytest.mark.ui
def test_canonical_point_pending_value_requires_finite_apply(tmp_path: Path) -> None:
    controller, _ = _controller(InMemoryMaterialRepository(), with_project=False)
    template = material_import_template("csv")
    source = tmp_path / template.filename
    source.write_bytes(template.data)
    controller.importTable(_file_url(source))
    controller.invalidateEditorInput("canonical:bh-25c:0", "Invalid canonical point.")
    before = controller.points[0]

    assert controller.canSave is False
    assert controller.setCanonicalPoint("bh-25c", 0, float("nan"), 0.15) is False
    assert controller.points[0] == before
    assert controller.canSave is False
    assert controller.setCanonicalPoint("bh-25c", 0, 12.5, 0.15) is True
    assert controller.points[0]["x"] == 12.5
    assert controller.points[0]["y"] == 0.15
    assert controller.canSave is True


@pytest.mark.ui
def test_controller_adds_table_and_image_series_then_removes_nonfinal(
    tmp_path: Path,
) -> None:
    controller, _ = _controller(InMemoryMaterialRepository(), with_project=False)
    controller.importSourceImage(_file_url(_IMAGE), 0)
    rendered = render_material_source(_IMAGE.name, _IMAGE.read_bytes())
    controller.setXAxis("linear", 0.0, 0.0, float(rendered.width_px), 100.0)
    controller.setYAxis(
        "linear", float(rendered.height_px), 0.0, 0.0, 0.5
    )
    controller.addPixelPoint(0.0, float(rendered.height_px))
    controller.addPixelPoint(float(rendered.width_px), 0.0)
    controller.createImageDraft("Example", "Ferrite", "Series", "Manual graph")

    assert controller.addTableSeries(
        "loss-extra",
        "loss-table",
        "T",
        "W/m3",
        100000.0,
        25.0,
        float("nan"),
        [{"x": 0.1, "y": 1000.0}, {"x": 0.2, "y": 4000.0}],
    ) is True
    assert controller.addImageSeries(
        "bh-second",
        "bh-curve",
        "A/m",
        "T",
        float("nan"),
        100.0,
        float("nan"),
    ) is True
    assert [item["seriesId"] for item in controller.series] == [
        "bh-manual",
        "loss-extra",
        "bh-second",
    ]
    assert controller.removeSeries("loss-extra") is True
    assert [item["seriesId"] for item in controller.series] == [
        "bh-manual",
        "bh-second",
    ]
    assert controller.removeSeries("bh-manual") is True
    assert controller.removeSeries("bh-second") is False
    assert "final" in controller.statusMessage


@pytest.mark.ui
def test_source_and_current_points_and_fit_attribution_are_truthful(tmp_path: Path) -> None:
    controller, _ = _controller(InMemoryMaterialRepository(), with_project=False)
    template = material_import_template("csv")
    source = tmp_path / template.filename
    source.write_bytes(template.data)
    controller.importTable(_file_url(source))

    original_source = controller.sourcePoints
    assert original_source == controller.points
    assert controller.sourceComparisonAvailable is True
    assert controller.fit["lossSeriesIds"] == ["loss-100khz", "loss-200khz"]

    point = controller.points[1]
    assert controller.setCanonicalPoint(
        str(point["seriesId"]), int(point["index"]), 123.0, 0.25
    ) is True
    assert controller.sourcePoints == original_source
    assert controller.points != original_source
    controller.saveDraft()
    revision_id = str(controller.selectedRevision["revisionId"])
    selected = controller.selectedMaterial
    controller.selectMaterial(
        str(selected["manufacturer"]),
        str(selected["name"]),
        str(selected["grade"]),
    )
    controller.selectRevision(revision_id)
    assert controller.sourceComparisonAvailable is False
    assert controller.sourcePoints == []
